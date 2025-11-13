import jwt
import requests
from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import api_view, APIView
import os
import logging
from rest_framework.generics import ListAPIView
from .serializers import *
import re
import io
from django.contrib.auth.models import User
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.utils.timezone import now
import logging
from collections import defaultdict
from django.utils.timezone import localtime, make_aware
from urllib.parse import quote
from django.core.signing import TimestampSigner, BadSignature, SignatureExpired
from django.utils.crypto import get_random_string
from io import BytesIO
from openpyxl import Workbook
from datetime import date
import datetime
from django.db import transaction
import json
from django_rest_passwordreset.signals import reset_password_token_created
from django.views.decorators.csrf import csrf_exempt

from django.shortcuts import get_object_or_404
from rest_framework.filters import SearchFilter
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from django.core.files.storage import FileSystemStorage
from datetime import datetime,timedelta
from django.http import HttpResponse,JsonResponse
from django.db.models import Count
from rest_framework.exceptions import NotFound
import calendar
import stripe
import zipfile
from io import BytesIO
import datetime
from django.contrib.auth import authenticate
from rest_framework_simplejwt.tokens import RefreshToken
from django.conf import settings
from django.contrib.auth.hashers import make_password
from .models import *
from django.core.mail import send_mail
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.contrib.auth import get_user_model
from .custom_tokens import CustomPasswordResetTokenGenerator 
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.parsers import JSONParser
from rest_framework.response import Response
from rest_framework import status
logger = logging.getLogger(__name__)
stripe.api_key = settings.STRIPE_SECRET_KEY
ALLOWED_FILE_EXTENSIONS = ['png', 'jpg', 'jpeg', 'gif']
from .models import Allergens

from django.utils import timezone
from datetime import timedelta
import jwt
import requests
from jwt.algorithms import RSAAlgorithm
from rest_framework.response import Response

APPLE_KEYS_URL = "https://appleid.apple.com/auth/keys"
APPLE_AUDIENCE = "app.raftersfoodservices.ie"

def verify_apple_token(identity_token: str):
    res = requests.get(APPLE_KEYS_URL)
    keys = res.json().get("keys", [])

    header = jwt.get_unverified_header(identity_token)
    key = next((k for k in keys if k["kid"] == header["kid"]), None)
    if not key:
        raise Exception("Apple public key not found")

    public_key = RSAAlgorithm.from_jwk(key)

    decoded = jwt.decode(
        identity_token,
        public_key,
        algorithms=["RS256"],
        audience=APPLE_AUDIENCE,
        issuer="https://appleid.apple.com",
    )
    return decoded


@api_view(["POST"])
def register(request):
    login_method = request.data.get("login_method", "email")
    email = request.data.get("email")
    username = request.data.get("username")
    password = request.data.get("password")
    confirm_password = request.data.get("confirm_password") or request.data.get("password_confirmation")
    school_type = request.data.get("school_type")
    school_id = request.data.get("school_id")  # frontend sends school_id

    if login_method in ["manager", "worker"]:
        if not username:
            return Response({"error": "Username is required."}, status=400)
        if not password or not confirm_password:
            return Response({"error": "Password and confirm password are required."}, status=400)
        if password != confirm_password:
            return Response({"error": "Passwords do not match."}, status=400)

        # Map school_id to correct field
        primary_school_id = school_id if school_type == "primary" else None
        secondary_school_id = school_id if school_type == "secondary" else None

        # MANAGER ACCOUNT
        if login_method == "manager":
            if Manager.objects.filter(username=username).exists():
                return Response({"error": "Username already exists."}, status=400)
            if password != confirm_password:
                return Response({"error": "Passwords do not match."}, status=400)


            primary_school_id = school_id if school_type == "primary" else None
            secondary_school_id = school_id if school_type == "secondary" else None
            manager = Manager.objects.create(
                username=username,
                password=make_password(password),
                school_type=school_type,
                primary_school_id=primary_school_id,
                secondary_school_id=secondary_school_id,
            )

            # Optional: return school name in response
            school_name = None
            if school_type == "primary" and primary_school_id:
                school_name = PrimarySchool.objects.filter(id=primary_school_id).first()
                school_name = school_name.school_name if school_name else "Unknown School"
            elif school_type == "secondary" and secondary_school_id:
                school_name = SecondarySchool.objects.filter(id=secondary_school_id).first()
                school_name = school_name.secondary_school_name if school_name else "Unknown School"

            return Response(
                {
                    "message": "Manager account created successfully.",
                    "user_type": "manager",
                    "id": manager.id,
                    "school_name": school_name or "Unknown School",
                    "school_type": school_type,
                    "primary_school": primary_school_id,
                    "secondary_school": secondary_school_id,
                },
                status=201,
            )

        # WORKER ACCOUNT

    if login_method == "worker":
        if Worker.objects.filter(username=username).exists():
            return Response({"error": "Username already exists."}, status=400)

        # Passwords already validated
        worker = Worker.objects.create(
            username=username,
            password=make_password(password)
        )

        return Response(
            {
                "message": "Worker account created successfully.",
                "user_type": "worker",
                "id": worker.id,
            },
            status=201,
        )

    # -------------------------------------------------
    # EMAIL-BASED REGISTRATION (Parent, Staff, Student)
    # -------------------------------------------------
    if login_method not in ["manager", "worker"]:
        if not email:
            return Response({"error": "Email is required for this registration type."}, status=400)

        # Check if email already registered
        if (
            StaffRegisteration.objects.filter(email=email).exists()
            or ParentRegisteration.objects.filter(email=email).exists()
            or SecondaryStudent.objects.filter(email=email).exists()
            or CanteenStaff.objects.filter(email=email).exists()
        ):
            return Response({"error": "Email already registered."}, status=400)

        existing_unverified = UnverifiedUser.objects.filter(email=email).first()
        if existing_unverified:
            if login_method == "email" and existing_unverified.created_at > timezone.now() - timedelta(seconds=60):
                return Response(
                    {"error": "A verification email has already been sent recently. Please check your inbox."},
                    status=400,
                )
            existing_unverified.delete()

        new_unverified = UnverifiedUser.objects.create(
            email=email,
            data=request.data,
            login_method=login_method,
        )

        if login_method == "email":
            verification_link = f"{settings.FRONTEND_URL}/verify-email/{new_unverified.token}/"
            send_mail(
                subject="Action Required: Confirm Your Email Address",
                message=f"""
Hi there,

Thank you for registering with us.

To complete your registration, please confirm your email address by clicking the link below:

{verification_link}

If you did not create an account with us, you can safely ignore this email.

Best regards,  
The Rafters Team
""",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
            )
            return Response(
                {"message": "Verification email sent. Please check your inbox."},
                status=200,
            )

        return Response(
            {
                "message": "OAuth login successful. Additional info required.",
                "token": str(new_unverified.token),
                "provider": login_method,
            },
            status=200,
        )

    return Response({"error": "Invalid registration type."}, status=400)




# Using your custom token generator
custom_token_generator = CustomPasswordResetTokenGenerator()

custom_token_generator = CustomPasswordResetTokenGenerator()
@api_view(["GET"])
def verify_email(request, token):
    try:
        unverified = UnverifiedUser.objects.get(token=token)

        # If already verified
        if unverified.is_verified:
            return Response(
                {"message": "This email is already verified."},
                status=status.HTTP_200_OK,
            )

        # Token expired (after 1 hour)
        if timezone.now() > unverified.created_at + timedelta(hours=1):
            old_data = unverified.data
            email = old_data.get("email")

            # Delete old record
            unverified.delete()

            # Generate a new token and resend
            new_unverified = UnverifiedUser.objects.create(
                email=email,
                data=old_data,
            )

            new_link = f"{settings.FRONTEND_URL}/verify-email/{new_unverified.token}/"
            send_mail(
                subject="New Verification Link - Rafters",
                message=f"""Hi there,

Your previous verification link has expired. Please use the new one below:

{new_link}

This link is valid for 1 hour.

Best regards,
The Rafters Team
""",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
            )

            return Response(
                {
                    "error": "Verification link has expired. A new one has been sent to your email."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        user_data = unverified.data
        user_type = user_data.get("user_type")

        phone_no = user_data.get("phone_no")
        if not phone_no or not str(phone_no).isdigit():
            user_data["phone_no"] = None  

        serializer = None

        # Handle user creation based on type
        if user_data.get("school_type"):
            if user_data["school_type"] == "primary":
                school = PrimarySchool.objects.get(id=user_data["school_id"])
            else:
                school = SecondarySchool.objects.get(id=user_data["school_id"])

            if user_type == "student":
                user_data["school"] = school.id
                serializer = SecondaryStudentSerializer(data=user_data)

            elif user_type == "staff":
                key = "primary_school" if user_data["school_type"] == "primary" else "secondary_school"
                user_data[key] = school.id
                serializer = StaffRegisterationSerializer(data=user_data)

            elif user_type == "canteenstaff":
                key = "primary_school" if user_data["school_type"] == "primary" else "secondary_school"
                user_data[key] = school.id
                serializer = CanteenStaffSerializer(data=user_data)

            else:
                return Response(
                    {"error": "Invalid user type."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        elif user_type == "parent":
            serializer = ParentRegisterationSerializer(data=user_data)

        else:
            return Response(
                {"error": "Invalid user type."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Save if valid
        if serializer.is_valid():
            serializer.save()
            unverified.is_verified = True
            unverified.save()

            return Response(
                {
                    "message": "Email verified successfully. User account has been created."
                },
                status=status.HTTP_201_CREATED,
            )
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    except UnverifiedUser.DoesNotExist:
        return Response(
            {"error": "Invalid or non-existing verification link."},
            status=status.HTTP_400_BAD_REQUEST,
        )



@api_view(["POST"])
def password_reset(request):
    email = request.data.get('email')
    if not email:
        return Response({"error": "Email is required."}, status=400)

    user = None
    for model in [ParentRegisteration, StaffRegisteration, SecondaryStudent, CanteenStaff]:
        try:
            user = model.objects.get(email=email)
            break
        except model.DoesNotExist:
            continue

    if not user:
        return Response({"error": "User not found."}, status=400)

    signer = TimestampSigner()
    signed_token = signer.sign(user.email)
    encoded_token = quote(signed_token)

    frontend_url = os.getenv('FRONTEND_URL', 'https://www.raftersfoodservices.ie/')
    reset_link = f'{frontend_url}/password-reset?token={encoded_token}'

    from_email = os.getenv('DEFAULT_FROM_EMAIL', 'support@raftersfoodservices.ie')

    send_mail(
        subject='Password Reset Request',
        message=f'Click the following link to reset your Rafters Food Services account password: {reset_link}',
        from_email=from_email,
        recipient_list=[user.email],
        fail_silently=False,
    )

    return Response({"message": "Password reset email sent."}, status=200)

@api_view(["POST"])
def password_reset_confirm(request):
    token = request.data.get("token")
    password = request.data.get("password")

    if not token or not password:
        return Response({"error": "Token and password must be provided."}, status=400)

    try:
        signer = TimestampSigner()
        email = signer.unsign(token, max_age=3600)  # 1 hour expiry
    except (BadSignature, SignatureExpired):
        return Response({"error": "Invalid or expired token."}, status=400)

    # Find user in all models
    user = None
    user_model = None
    for model in [ParentRegisteration, StaffRegisteration, SecondaryStudent, CanteenStaff]:
        try:
            user = model.objects.get(email=email)
            user_model = model
            break
        except model.DoesNotExist:
            continue

    if not user:
        return Response({"error": "User not found."}, status=400)

    user_model.objects.filter(email=email).update(password=make_password(password))
    updated_user = user_model.objects.get(email=email)

    if not check_password(password, updated_user.password):
        return Response({"error": "Password was not set correctly."}, status=500)

    return Response({"message": "Password reset successful."}, status=200)
@api_view(["POST"])
def login(request):

    serializer = LoginSerializer(data=request.data)
    
    if serializer.is_valid():
        user = serializer.validated_data['user']
        user_type = serializer.validated_data['user_type'] 
       
        refresh = RefreshToken.for_user(user)
        access_token = str(refresh.access_token)
        refresh_token = str(refresh)
        request.session['user_id'] = user.id
        request.session['user_email'] = user.email
        request.session['user_type'] = user_type
 
        return Response({
            'access': access_token,
            'refresh': refresh_token,
            'user_type': user_type,
            'user_id': user.id,
            'message': 'Login successful'
        }, status=status.HTTP_200_OK)
    
    return Response({
        'detail': 'Invalid credentials'
    }, status=status.HTTP_401_UNAUTHORIZED)



@api_view(['POST',])

def admin_login(request):
     username=request.data.get("username")
     password=request.data.get("password")
     if not username or not password:
        return Response({'detail': 'Both username and password are required.'}, status=status.HTTP_400_BAD_REQUEST)

     user = authenticate(username=username, password=password)

     if user is None:
        return Response({'detail': 'Invalid username or password.'}, status=status.HTTP_401_UNAUTHORIZED)

     if not user.is_staff:
        return Response({'detail': 'You do not have permission to access this resource.'}, status=status.HTTP_403_FORBIDDEN)

     return Response({'detail': 'Login successful!'}, status=status.HTTP_200_OK)




@api_view(["GET"])
def get_user_info(request, user_type, id):
    if user_type == "parent":
        try:
            parent = ParentRegisteration.objects.get(id=id)
            students = PrimaryStudentsRegister.objects.filter(parent=parent)
            parent_serializer = ParentRegisterationSerializer(parent)
            student_serializer = PrimaryStudentSerializer(students, many=True)

            parent_data = parent_serializer.data
            parent_data.pop('password', None)

            student_data = student_serializer.data
            for student in student_data:
                student.pop('password', None)

            
            if student_data:
           
                school_type = "primary" if students[0].school else None
                parent_data['school_type'] = school_type
            else:
                parent_data['school_type'] = None

            return Response({
                "user_type": "parent",
                "user_id": id,
                "parent": parent_data,
                "students": student_data
            }, status=status.HTTP_200_OK)

        except ParentRegisteration.DoesNotExist:
            return Response({"error": "Parent not found."}, status=status.HTTP_404_NOT_FOUND)

    elif user_type == "staff":
        try:
            staff = StaffRegisteration.objects.get(id=id)
            students = PrimaryStudentsRegister.objects.filter(staff=staff)
            staff_serializer = StaffRegisterationSerializer(staff)
            student_serializer = PrimaryStudentSerializer(students, many=True)

            staff_data = staff_serializer.data
            staff_data.pop('password', None)

            student_data = student_serializer.data
            for student in student_data:
                student.pop('password', None)

            # Adding school type to the staff response
            school_type = "primary" if staff.primary_school else "secondary" if staff.secondary_school else None
            staff_data['school_type'] = school_type

            return Response({
                "user_type": "staff",
                "user_id": id,
                "staff": staff_data,
                "students": student_data
            }, status=status.HTTP_200_OK)

        except StaffRegisteration.DoesNotExist:
            return Response({"error": "Staff not found."}, status=status.HTTP_404_NOT_FOUND)

    elif user_type == "student":
        try:
            student = SecondaryStudent.objects.get(id=id)
            student_serializer = SecondaryStudentSerializer(student)

            student_data = student_serializer.data
            student_data.pop('password', None)

            # Adding school type to the student response
            school_type = "secondary" if student.school else None
            student_data['school_type'] = school_type

            return Response({
                "user_type": "student",
                "user_id": id,
                "student": student_data
            }, status=status.HTTP_200_OK)

        except SecondaryStudent.DoesNotExist:
            return Response({"error": "Student not found."}, status=status.HTTP_404_NOT_FOUND)

    else:
        return Response({"error": "Invalid user_type."}, status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET", "PUT","DELETE"])
def update_user_info(request, user_type, id):
    # Handling GET request to fetch user data
    if request.method == "GET":
        if user_type == "parent":
            try:
                parent = ParentRegisteration.objects.get(id=id)
                parent_serializer = ParentRegisterationSerializer(parent)
                parent_data = parent_serializer.data 
                if 'password' in parent_data:
                    parent_data.pop('password') 
                return Response({
                    "user_type": "parent",
                    "user_id": id,
                    "parent": parent_data
                }, status=status.HTTP_200_OK)
            except ParentRegisteration.DoesNotExist:
                return Response({"error": "Parent not found."}, status=status.HTTP_404_NOT_FOUND)

        elif user_type == "staff":
            try:
                staff = StaffRegisteration.objects.get(id=id)
                staff_serializer = StaffRegisterationSerializer(staff)
                staff_data = staff_serializer.data 
                if 'password' in staff_data:
                    staff_data.pop('password')  
                return Response({
                    "user_type": "staff",
                    "user_id": id,
                    "staff": staff_data
                }, status=status.HTTP_200_OK)
            except StaffRegisteration.DoesNotExist:
                return Response({"error": "Staff not found."}, status=status.HTTP_404_NOT_FOUND)

        elif user_type == "student":
            try:
                student = SecondaryStudent.objects.get(id=id)
                student_serializer = SecondaryStudentSerializer(student)
                student_data = student_serializer.data  
                if 'password' in student_data:
                    student_data.pop('password') 
                return Response({
                    "user_type": "student",
                    "user_id": id,
                    "student": student_data
                }, status=status.HTTP_200_OK)
            except SecondaryStudent.DoesNotExist:
                return Response({"error": "Student not found."}, status=status.HTTP_404_NOT_FOUND)

        else:
            return Response({"error": "Invalid user_type."}, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == "PUT":
        if user_type == "parent":
            try:
                parent = ParentRegisteration.objects.get(id=id)

             
                if 'password' not in request.data:
                    request.data['password'] = parent.password  

               
                serializer = ParentRegisterationSerializer(parent, data=request.data, partial=True)
                if serializer.is_valid():
                    updated_parent = serializer.save()

                
                    updated_data = serializer.data
                    if 'password' in updated_data:
                        updated_data.pop('password')

                    return Response({
                        "user_type": "parent",
                        "user_id": id,
                        "parent": updated_data,
                        "message": "Profile updated successfully."
                    }, status=status.HTTP_200_OK)
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            except ParentRegisteration.DoesNotExist:
                return Response({"error": "Parent not found."}, status=status.HTTP_404_NOT_FOUND)

        elif user_type == "staff":
            try:
                staff = StaffRegisteration.objects.get(id=id)

             
                if 'password' not in request.data:
                    request.data['password'] = staff.password 

               
                serializer = StaffRegisterationSerializer(staff, data=request.data, partial=True)
                if serializer.is_valid():
                    updated_staff = serializer.save()

                    
                    updated_data = serializer.data
                    if 'password' in updated_data:
                        updated_data.pop('password')

                    return Response({
                        "user_type": "staff",
                        "user_id": id,
                        "staff": updated_data,
                        "message": "Profile updated successfully."
                    }, status=status.HTTP_200_OK)
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            except StaffRegisteration.DoesNotExist:
                return Response({"error": "Staff not found."}, status=status.HTTP_404_NOT_FOUND)

        elif user_type == "student":
            try:
                student = SecondaryStudent.objects.get(id=id)

              
                if 'password' not in request.data:
                    request.data['password'] = student.password 
              
                serializer = SecondaryStudentSerializer(student, data=request.data, partial=True)
                if serializer.is_valid():
                    updated_student = serializer.save()

                    
                    updated_data = serializer.data
                    if 'password' in updated_data:
                        updated_data.pop('password')

                    return Response({
                        "user_type": "student",
                        "user_id": id,
                        "student": updated_data,
                        "message": "Profile updated successfully."
                    }, status=status.HTTP_200_OK)
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            except SecondaryStudent.DoesNotExist:
                return Response({"error": "Student not found."}, status=status.HTTP_404_NOT_FOUND)

        else:
            return Response({"error": "Invalid user_type."}, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == "DELETE":
        if user_type == "parent":
            try:
                parent = ParentRegisteration.objects.get(id=id)
                parent.delete()
                return Response({
                    "message": "Parent deleted successfully."
                }, status=status.HTTP_204_NO_CONTENT)
            except ParentRegisteration.DoesNotExist:
                return Response({"error": "Parent not found."}, status=status.HTTP_404_NOT_FOUND)

        elif user_type == "staff":
            try:
                staff = StaffRegisteration.objects.get(id=id)
                staff.delete()
                return Response({
                    "message": "Staff deleted successfully."
                }, status=status.HTTP_204_NO_CONTENT)
            except StaffRegisteration.DoesNotExist:
                return Response({"error": "Staff not found."}, status=status.HTTP_404_NOT_FOUND)

        elif user_type == "student":
            try:
                student = SecondaryStudent.objects.get(id=id)
                student.delete()
                return Response({
                    "message": "Student deleted successfully."
                }, status=status.HTTP_204_NO_CONTENT)
            except SecondaryStudent.DoesNotExist:
                return Response({"error": "Student not found."}, status=status.HTTP_404_NOT_FOUND)

        else:
            return Response({"error": "Invalid user_type."}, status=status.HTTP_400_BAD_REQUEST)
# Canteen Staff
@csrf_exempt
@api_view(['GET',])
def get_cateenstaff(request):
   if request.method == "GET":
        staff = CanteenStaff.objects.all()
        serializer = CanteenStaffSerializer(staff,many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET', 'PUT', 'DELETE'])
def cateenstaff_by_id(request, pk):
    try:
        staff = CanteenStaff.objects.get(pk=pk)
    except CanteenStaff.DoesNotExist:
        raise NotFound({'error': 'Staff not found'})

    if request.method == 'GET':
        serializer = CanteenStaffSerializer(staff)
        data = serializer.data  # Serialized data

        # Add school_name field dynamically
        if staff.school_type == 'primary' and staff.primary_school:
            data['school_name'] = staff.primary_school.school_name
        elif staff.school_type == 'secondary' and staff.secondary_school:
            data['school_name'] = staff.secondary_school.secondary_school_name
        else:
            data['school_name'] = 'Unknown School'

        return Response(data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = CanteenStaffSerializer(staff, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        staff.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

@csrf_exempt
@api_view(['GET'])
def get_managers(request):
    managers = Manager.objects.all()
    serializer = ManagerSerializer(managers, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['GET', 'PUT', 'DELETE'])
def manager_by_id(request, pk):
    try:
        manager = Manager.objects.get(pk=pk)
    except Manager.DoesNotExist:
        raise NotFound({'error': 'Manager not found'})

    if request.method == 'GET':
        serializer = ManagerSerializer(manager)
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = ManagerSerializer(manager, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        manager.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ---------------------------
# WORKER APIS
# ---------------------------

@csrf_exempt
@api_view(['GET'])
def get_workers(request):
    workers = Worker.objects.all()
    serializer = WorkerSerializer(workers, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['GET', 'PUT', 'DELETE'])
def worker_by_id(request, pk):
    try:
        worker = Worker.objects.get(pk=pk)
    except Worker.DoesNotExist:
        raise NotFound({'error': 'Worker not found'})

    if request.method == 'GET':
        serializer = WorkerSerializer(worker)
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = WorkerSerializer(worker, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        worker.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

def generate_unique_username(first_name, last_name):
    from admin_section.models import PrimaryStudentsRegister
    base_username = f"{first_name.lower()}_{last_name.lower()}"
    username = base_username
    counter = 1

    while PrimaryStudentsRegister.objects.filter(username=username).exists():
        username = f"{base_username}_{counter}"
        counter += 1

    return username


@api_view(['POST'])
def add_child(request):
    school_name = request.data.get('school')
    first_name = request.data.get('first_name')
    last_name = request.data.get('last_name')
    class_year = request.data.get('class_year')
    teacher = request.data.get('teacher')
    allergies_data = request.data.get('allergies', [])
    user_id = request.data.get('user_id')
    user_type = request.data.get('user_type')

    # Validation checks
    if not school_name:
        return Response({"error": "School name is required."}, status=status.HTTP_400_BAD_REQUEST)

    if not first_name:
        return Response({"error": "First name is required."}, status=status.HTTP_400_BAD_REQUEST)

    if not last_name:
        return Response({"error": "Last name is required."}, status=status.HTTP_400_BAD_REQUEST)

    if not class_year:
        return Response({"error": "Class year is required."}, status=status.HTTP_400_BAD_REQUEST)

    if not user_id:
        return Response({"error": "User ID is required."}, status=status.HTTP_400_BAD_REQUEST)

    if not user_type:
        return Response({"error": "User type is required."}, status=status.HTTP_400_BAD_REQUEST)

    
    try:
        school = PrimarySchool.objects.get(school_name=school_name)
    except PrimarySchool.DoesNotExist:
        return Response({"error": f"School '{school_name}' does not exist."}, status=status.HTTP_400_BAD_REQUEST)

    user = None
    if user_type == 'parent':
        try:
            user = ParentRegisteration.objects.get(id=user_id)
        except ParentRegisteration.DoesNotExist:
            return Response({"error": f"Parent with ID {user_id} does not exist."}, status=status.HTTP_400_BAD_REQUEST)
    elif user_type == 'staff':
        try:
            user = StaffRegisteration.objects.get(id=user_id)
        except StaffRegisteration.DoesNotExist:
            return Response({"error": f"Staff with ID {user_id} does not exist."}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({"error": "Invalid user type."}, status=status.HTTP_400_BAD_REQUEST)

    username = generate_unique_username(first_name, last_name)
 

    # Create student data
    student_data = {
        'first_name': first_name,
        'last_name': last_name,
        'username': username, 
        'class_year': class_year,
        'school': school.id,
        'allergies': allergies_data,
        'teacher': teacher
    }

  
    if user_type == 'parent':
        student_data['parent'] = user_id
    elif user_type == 'staff':
        student_data['staff'] = user_id

    
    serializer = PrimaryStudentSerializer(data=student_data)

    if serializer.is_valid():
        student = serializer.save()  
        return Response({
            "message": "Child added successfully",
            "student": serializer.data,
            "user_type": user_type,
            "user_id": user_id,
            "school_type": "primary"  
        }, status=status.HTTP_201_CREATED)
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


    
@api_view(['GET', 'PUT', 'DELETE'])
def edit_child(request, child_id):
    child = get_object_or_404(PrimaryStudentsRegister, id=child_id)

    if request.method == 'GET':
        serializer = PrimaryStudentSerializer(child)
        return Response({'child': serializer.data}, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        first_name = request.data.get('first_name', child.first_name)
        last_name = request.data.get('last_name', child.last_name)
        class_year = request.data.get('class_year', child.class_year)
        school_id = request.data.get('school', None)
        teacher_id = request.data.get('teacher_id', None)  
        allergies = request.data.get('allergies', [])


        if school_id:
            try:
                school = PrimarySchool.objects.get(id=school_id)
                child.school = school
            except PrimarySchool.DoesNotExist:
                return Response({'message': 'School not found'}, status=status.HTTP_400_BAD_REQUEST)

        if teacher_id:
            try:
                teacher = Teacher.objects.get(id=teacher_id)  
                child.teacher = teacher
            except Teacher.DoesNotExist:
                return Response({'message': 'Teacher not found'}, status=status.HTTP_400_BAD_REQUEST)

        # --- Update child basic info ---
        child.first_name = first_name
        child.last_name = last_name
        child.class_year = class_year

        # --- Update allergies ---
        if allergies is not None:
            child.allergies.clear()
            for allergy in allergies:
                allergen = Allergens.objects.filter(allergy=allergy).first()
                if allergen:
                    child.allergies.add(allergen)

        # --- Save child ---
        child.save()

        return Response({
            'message': 'Child details updated successfully.',
            'child': PrimaryStudentSerializer(child).data
        }, status=status.HTTP_200_OK)

    elif request.method == 'DELETE':
        child.delete()
        return Response({'message': 'Child record deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)


# Primary School Sction
@csrf_exempt
@api_view(['POST'])
def add_primary_school(request):
    if request.method=='POST':
        serializer=PrimarySchoolSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data,status=status.HTTP_201_CREATED, )
        else:
            return Response({ "error" : serializer.errors},status = status.HTTP_400_BAD_REQUEST)
        
@csrf_exempt
@api_view(['GET',])
def primary_school(request):
    if request.method == "GET":
        try:
           
            details = PrimarySchool.objects.annotate(student_count=Count('student'))
        except:
            return Response(status=status.HTTP_404_NOT_FOUND)
    
        serializer = PrimarySchoolSerializer(details, many=True)
    
        for school_data in serializer.data:
    
            school_data['student_count'] = school_data.get('student_count', 0)
        
        return Response(serializer.data)

@api_view(['GET', 'DELETE', 'PUT'])
def delete_primary_school(request, pk):
    try:
        school = PrimarySchool.objects.annotate(student_count=Count('student')).get(pk=pk)
    except PrimarySchool.DoesNotExist:
        return Response({"error": "PrimarySchool not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "PUT":
        serializer = PrimarySchoolSerializer(school, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_204_NO_CONTENT)
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)

    if request.method == "DELETE":
        school.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    serializer = PrimarySchoolSerializer(school)
    response_data = serializer.data
    response_data['student_count'] = school.student_count
    return Response(response_data, status=status.HTTP_200_OK)

class PrimarySearch(ListAPIView):

   queryset=PrimarySchool.objects.all()
   serializer_class=PrimarySchoolSerializer
   filter_backends=[SearchFilter]
   search_fields=['school_name','school_email']
   

# For Teacher
@api_view(['GET', 'POST'])
def add_and_get_teacher(request, school_id):
    school = get_object_or_404(PrimarySchool, pk=school_id)
    if request.method == 'GET':
        teachers = Teacher.objects.filter(school=school)
        teacher_serializer = TeacherSerializer(teachers, many=True)
        return Response(teacher_serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'POST':
        teacher_name = request.data.get('teacher_name')
        class_year = request.data.get('class_year')

        new_teacher = Teacher(
            teacher_name=teacher_name,
            class_year=class_year,
            school=school
        )
        new_teacher.save()

        teacher_serializer = TeacherSerializer(new_teacher)
        return Response(teacher_serializer.data, status=status.HTTP_201_CREATED)

@api_view(['GET', 'PUT', 'DELETE'])
def update_delete_teacher(request, school_id, teacher_id):
    school = get_object_or_404(PrimarySchool, pk=school_id)
    teacher = get_object_or_404(Teacher, pk=teacher_id, school=school)

    if request.method == 'GET':
        teacher_serializer = TeacherSerializer(teacher)
        return Response(teacher_serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        teacher_name = request.data.get('teacher_name', teacher.teacher_name)
        class_year = request.data.get('class_year', teacher.class_year)

        teacher.teacher_name = teacher_name
        teacher.class_year = class_year
        teacher.save()

        teacher_serializer = TeacherSerializer(teacher)
        return Response(teacher_serializer.data, status=status.HTTP_200_OK)
    elif request.method == 'DELETE':
        teacher.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['POST'])
def get_teachers(request):
 
    school_id = request.data.get('school_id')
    class_year = request.data.get('class_year')

    if not school_id or not class_year:
        return Response({'error': 'Both school_id and class_year are required.'}, status=status.HTTP_400_BAD_REQUEST)

   
    teachers = Teacher.objects.filter(school_id=school_id, class_year=class_year)

    if not teachers.exists():
        return Response({'error': 'No teachers found for the given school and class year.'}, status=status.HTTP_404_NOT_FOUND)

    teacher_details = []

    for teacher in teachers:
        teacher_data = {
            'id': teacher.id,
            'teacher_name': teacher.teacher_name,
            'class_year': teacher.class_year,
            'school_id': teacher.school.id,
            'school_name': teacher.school.school_name  
        }
        teacher_details.append(teacher_data)

 
    return Response({
        'message': 'Teachers retrieved successfully!',
        'teachers': teacher_details
    }, status=status.HTTP_200_OK)

@api_view(['GET'])
def list_all_teachers(request):
    teachers = Teacher.objects.all()
    serializer = TeacherSerializer(teachers, many=True)
    return Response({'teachers': serializer.data}, status=status.HTTP_200_OK)


@api_view(['GET', 'PUT', 'DELETE'])
def teacher_detail(request, teacher_id):
    try:
        teacher = Teacher.objects.get(id=teacher_id)
    except Teacher.DoesNotExist:
        return Response({'error': 'Teacher not found.'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = TeacherSerializer(teacher)
        return Response(serializer.data)

    elif request.method == 'PUT':
        serializer = TeacherSerializer(teacher, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Teacher updated successfully.', 'data': serializer.data})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        teacher.delete()
        return Response({'message': 'Teacher deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)

# For Primary School Student


@api_view(['GET', 'POST'])
def get_student_detail(request, school_id):
  
    school = get_object_or_404(PrimarySchool, pk=school_id)
   
    if request.method == 'GET':
      
        students = PrimaryStudentsRegister.objects.filter(school=school)
        student_serializer = PrimaryStudentSerializer(students, many=True)
        return Response(student_serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'POST':
     
        first_name = request.data.get('first_name')
        last_name = request.data.get('last_name')
        class_year = request.data.get('class_year')
        teacher_id = request.data.get('teacher_id')

      
        teacher = get_object_or_404(Teacher, pk=teacher_id)

     
        new_student = PrimaryStudentsRegister(
            first_name=first_name,
            last_name=last_name,
            class_year=class_year,
            school=school,
            teacher=teacher
        )
        
      
        new_student.save()

       
        student_serializer = PrimaryStudentSerializer(new_student)
        return Response(student_serializer.data, status=status.HTTP_201_CREATED)

@api_view(['GET', 'PUT', 'DELETE'])
def update_delete_student(request, school_id, student_id):
    school = get_object_or_404(PrimarySchool, pk=school_id)
    student = get_object_or_404(PrimaryStudentsRegister, pk=student_id, school=school)

    def exclude_password(data):
        if 'password' in data:
            data.pop('password')
        return data

   
    if request.method == 'GET':
        student_serializer = PrimaryStudentSerializer(student)
        student_data = exclude_password(student_serializer.data)  
        return Response(student_data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
      
        first_name = request.data.get('first_name', student.first_name)
        last_name = request.data.get('last_name', student.last_name)
        class_year = request.data.get('class_year', student.class_year)
        teacher_id = request.data.get('teacher', None)

     
        if teacher_id:
            teacher = get_object_or_404(Teacher, pk=teacher_id)
            student.teacher = teacher

        student.first_name = first_name
        student.last_name = last_name
        student.class_year = class_year

        
        password = request.data.get('password', None)
        if password:
            student.set_password(password)

       
        student.save()

        student_serializer = PrimaryStudentSerializer(student)
        student_data = exclude_password(student_serializer.data)
        return Response(student_data, status=status.HTTP_200_OK)

    elif request.method == 'DELETE':
        student.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)     

    elif request.method == 'DELETE':
        student.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

class StudentSearch(ListAPIView):
   queryset=PrimaryStudentsRegister.objects.all()
   serializer_class=PrimaryStudentsRegister
   filter_backends=[SearchFilter]
   search_fields=['student_name','class_year',"teacher"]


# Secondary School Section

@csrf_exempt
@api_view(['POST'])
def add_secondary_school(request):
    if request.method=='POST':
        serializer=SecondarySchoolSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data,status=status.HTTP_201_CREATED, )
        else:
            return Response({ "error" : serializer.errors},status = status.HTTP_400_BAD_REQUEST)
        
@csrf_exempt
@api_view(['GET',])
def secondary_school(request):
   if request.method == "GET":
        try:
           
            details = SecondarySchool.objects.annotate(student_count=Count('student'))
        except:
            return Response(status=status.HTTP_404_NOT_FOUND)
    
        serializer = SecondarySchoolSerializer(details, many=True)
    
        for school_data in serializer.data:
    
            school_data['student_count'] = school_data.get('student_count', 0)
        
        return Response(serializer.data)


@api_view(['GET','DELETE','PUT'])
def delete_secondary_school(request,pk):
    try:
        school = SecondarySchool.objects.annotate(student_count=Count('student')).get(pk=pk)
    except SecondarySchool.DoesNotExist:
        return Response({"error": "SecondarySchool not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "PUT":
        serializer = SecondarySchoolSerializer(school, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_204_NO_CONTENT)
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)

    if request.method == "DELETE":
        school.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    serializer = SecondarySchoolSerializer(school)
    response_data = serializer.data
    response_data['student_count'] = school.student_count
    return Response(response_data, status=status.HTTP_200_OK)


# For Secondary Student

@csrf_exempt
@api_view(['POST'])
def add_secondary_student(request, school_id):
    if request.method == 'POST':
        try:
            school = SecondarySchool.objects.get(pk=school_id)
        except SecondarySchool.DoesNotExist:
            return Response({"error": "Secondary school not found."}, status=status.HTTP_404_NOT_FOUND)

        request.data['secondary_school'] = school.id 
        serializer = SecondaryStudentSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response({"error": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT', 'DELETE'])
def update_delete_secondary_student(request, school_id, student_id):
 
    school = get_object_or_404(SecondarySchool, pk=school_id)
    student = get_object_or_404(SecondaryStudent, pk=student_id, school=school)

   
    if request.method == 'GET':
        student_serializer = SecondaryStudentSerializer(student)
       
        student_data = student_serializer.data
        student_data.pop('password', None)  
        return Response(student_data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        student_name = request.data.get('secondary_student_name', student.secondary_student_name)
        class_year = request.data.get('secondary_class_year', student.secondary_class_year)

       
        student.secondary_student_name = student_name
        student.secondary_class_year = class_year
        student.save()

        student_serializer = SecondaryStudentSerializer(student)
        student_data = student_serializer.data
        student_data.pop('password', None)  
        return Response(student_data, status=status.HTTP_200_OK)

    
    elif request.method == 'DELETE':
        student.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
@api_view(['GET', 'PUT', 'DELETE'])
def update_delete_secondary_student(request, school_id, student_id):
    def exclude_password(data):
        if 'password' in data:
            data.pop('password')
        return data
    school = get_object_or_404(SecondarySchool, pk=school_id)
    student = get_object_or_404(SecondaryStudent, pk=student_id, school=school)

    if request.method == 'GET':
        student_serializer = SecondaryStudentSerializer(student)
        return Response(student_serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
       
        first_name = request.data.get('first_name', student.first_name) 
        last_name = request.data.get('last_name', student.last_name)      
        class_year = request.data.get('secondary_class_year', student.class_year)

        
        student.first_name = first_name
        student.last_name = last_name
        student.class_year = class_year
        student.save()

        student_serializer = SecondaryStudentSerializer(student)
        return Response(student_serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'DELETE':
        student.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
# Category
@csrf_exempt
@api_view(['GET',])
def get_category(request):
   if request.method == "GET":
        category = Categories.objects.all()
        serializer = CategoriesSerializer(category,many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
#    Allergerns
@csrf_exempt
@api_view(['GET',])
def get_allergy(request):
   if request.method == "GET":
        allergy = Allergens.objects.all()
        serializer = AllergenSerializer(allergy,many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)



@api_view(['POST'])
def activate_cycle(request):
    school_id = request.data.get('school_id')
    school_type = request.data.get('school_type')
    cycle_name = request.data.get('cycle_name')

    if not school_id or not school_type or not cycle_name:
        return Response({'error': 'Missing required fields'}, status=status.HTTP_400_BAD_REQUEST)

    school_model = PrimarySchool if school_type == 'primary' else SecondarySchool
    school = school_model.objects.filter(id=school_id).first()
    if not school:
        return Response({'error': f'{school_type.capitalize()} School not found'}, status=status.HTTP_404_NOT_FOUND)

    # Get menus of this cycle
    menus = Menu.objects.filter(cycle_name=cycle_name)
    if not menus.exists():
        return Response({'error': f'No menus found for cycle "{cycle_name}".'}, status=status.HTTP_404_NOT_FOUND)

    # 🔴 First deactivate all existing menus linked with this school
    if school_type == 'primary':
        previous_menus = Menu.objects.filter(primary_schools=school, is_active=True)
        for menu in previous_menus:
            menu.primary_schools.remove(school)
            # If menu is not linked to any school, deactivate it globally
            if not menu.primary_schools.exists() and not menu.secondary_schools.exists():
                menu.is_active = False
            menu.save()
    else:
        previous_menus = Menu.objects.filter(secondary_schools=school, is_active=True)
        for menu in previous_menus:
            menu.secondary_schools.remove(school)
            if not menu.primary_schools.exists() and not menu.secondary_schools.exists():
                menu.is_active = False
            menu.save()

    # ✅ Now activate the new cycle menus
    updated_menus = []
    for menu in menus:
        menu.is_active = True
        menu.save()

        if school_type == 'primary':
            menu.primary_schools.add(school)
        else:
            menu.secondary_schools.add(school)

        updated_menus.append({
            'id': menu.id,
            'name': menu.name,
            'price': str(menu.price),
            'menu_day': menu.menu_day,
            'cycle_name': menu.cycle_name,
            'is_active': menu.is_active,
            'schools': {
                'primary': [s.school_name for s in menu.primary_schools.all()],
                'secondary': [s.secondary_school_name for s in menu.secondary_schools.all()]
            }
        })

    return Response({
        'message': f'Cycle "{cycle_name}" activated successfully for {school.school_name if school_type == "primary" else school.secondary_school_name}!',
        'menus': updated_menus
    }, status=status.HTTP_200_OK)


@api_view(['POST', 'GET', 'DELETE'])
def get_complete_menu(request):
    if request.method == 'POST':
        school_id = request.data.get('school_id')  
        school_type = request.data.get('school_type')  
        cycle_name = request.data.get('cycle_name')
        start_date = request.data.get('start_date')  
        end_date = request.data.get('end_date')  

  
        if not school_id or not school_type or not cycle_name:
            return Response({'error': 'school_id, school_type, and cycle_name are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else None
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else None
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        if school_type == 'primary':
            school = PrimarySchool.objects.filter(id=school_id).first() 
        elif school_type == 'secondary':
            school = SecondarySchool.objects.filter(id=school_id).first()
        else:
            return Response({'error': 'Invalid school_type. Use "primary" or "secondary".'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not school:
            return Response({'error': f'{school_type.capitalize()} school with ID {school_id} not found.'}, status=status.HTTP_404_NOT_FOUND)

        menus = Menu.objects.filter(
            primary_school=school if school_type == 'primary' else None,
            secondary_school=school if school_type == 'secondary' else None,
            cycle_name=cycle_name
        )

        if start_date and end_date:
            menus = menus.filter(start_date__lte=start_date, end_date__gte=end_date)

        if not menus.exists():
            return Response({'error': f'No menus found for cycle "{cycle_name}" in the specified school.'}, status=status.HTTP_404_NOT_FOUND)

        menu_data = {}
        for menu in menus:
            menu_data.setdefault(menu.menu_day, []).append({
                'id': menu.id,
                'name': menu.name,
                'price': menu.price,
                'category': menu.category.name_category,
                'menu_date': menu.menu_date,
                'cycle_name': menu.cycle_name,
                'is_active': menu.is_active 
            })

        return Response({'status': 'Menus retrieved successfully!', 'menus': menu_data}, status=status.HTTP_200_OK)

    elif request.method == 'DELETE':
        school_id = request.data.get('school_id')
        school_type = request.data.get('school_type')
        cycle_name = request.data.get('cycle_name')
        start_date = request.data.get('start_date')  
        end_date = request.data.get('end_date')  

        if not school_id or not school_type or not cycle_name:
            return Response({'error': 'school_id, school_type, and cycle_name are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else None
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else None
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        if school_type == 'primary':
            school = PrimarySchool.objects.filter(id=school_id).first()  
        elif school_type == 'secondary':
            school = SecondarySchool.objects.filter(id=school_id).first()
        else:
            return Response({'error': 'Invalid school_type. Use "primary" or "secondary".'}, status=status.HTTP_400_BAD_REQUEST)

        if not school:
            return Response({'error': f'{school_type.capitalize()} school with ID {school_id} not found.'}, status=status.HTTP_404_NOT_FOUND)

        menus_to_delete = Menu.objects.filter(
            primary_school=school if school_type == 'primary' else None,
            secondary_school=school if school_type == 'secondary' else None,
            cycle_name=cycle_name
        )

        if start_date and end_date:
            menus_to_delete = menus_to_delete.filter(start_date__lte=start_date, end_date__gte=end_date)

        if not menus_to_delete.exists():
            return Response({'error': f'No menus found for cycle "{cycle_name}" in the specified school.'}, status=status.HTTP_404_NOT_FOUND)

        menus_to_delete.delete()

        return Response({'message': f'All menus for cycle "{cycle_name}" in school {school_id} have been deleted.'}, status=status.HTTP_204_NO_CONTENT)

    return Response({'error': 'Invalid request method.'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)


@api_view(["POST"])
def get_active_menu(request):
    school_type = request.data.get('school_type')
    school_id = request.data.get('school_id')

    if not school_type or not school_id:
        return Response({"detail": "Both 'school_type' and 'school_id' must be provided."},
                        status=status.HTTP_400_BAD_REQUEST)

    try:
        if school_type == 'primary':
            menus = Menu.objects.filter(
                primary_schools__id=school_id   # ✅ updated for M2M
            ).select_related('category')
            
            subquery = Menu.objects.filter(
                primary_schools__id=school_id   # ✅ updated for M2M
            ).values('cycle_name').distinct()

        elif school_type == 'secondary':
            menus = Menu.objects.filter(
                secondary_schools__id=school_id   # ✅ updated for M2M
            ).select_related('category')
            
            subquery = Menu.objects.filter(
                secondary_schools__id=school_id   # ✅ updated for M2M
            ).values('cycle_name').distinct()

        else:
            return Response({"detail": "Invalid school type. Please provide either 'primary' or 'secondary'."},
                            status=status.HTTP_400_BAD_REQUEST)

        # filter only active
        active_menus = [menu for menu in menus if menu.is_active]

        # collect active cycles
        active_cycles = [
            {"cycle": cycle['cycle_name']} 
            for cycle in subquery 
            if any(menu.cycle_name == cycle['cycle_name'] for menu in active_menus)
        ]

        menus_data = MenuSerializer(active_menus, many=True).data

        weekly_menu = {
            "Monday": [],
            "Tuesday": [],
            "Wednesday": [],
            "Thursday": [],
            "Friday": [],
            "Saturday": [],
            "Sunday": []
        }

        for menu in menus_data:
            day_of_week = menu['menu_day']
            if day_of_week in weekly_menu:
                menu_items = MenuItems.objects.filter(item_name=menu['name'])
                menu_items_data = []

                for item in menu_items:
                    menu_items_data.append({
                        "item_name": item.item_name,
                        "item_description": item.item_description,
                        "ingredients": item.ingredients,
                        "nutrients": item.nutrients,
                        "allergies": [allergy.allergy for allergy in item.allergies.all()],
                        "image_url": request.build_absolute_uri(item.image.url) if item.image else None
                    })

                category_name = menu.get('category', None)  
                weekly_menu[day_of_week].append({
                    'id': menu['id'],
                    "name": menu['name'],
                    "price": menu['price'],
                    "menu_date": menu['menu_date'],
                    "cycle_name": menu['cycle_name'],
                    "category": category_name,  
                    "is_active": menu['is_active'],
                    "menu_items": menu_items_data
                })

        return Response({
            "menus": weekly_menu,
            "cycles": active_cycles
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(["GET", "PUT"])
def edit_menu(request, id):
    print(f"Received request to edit menu item with id: {id}")
    if request.method == 'GET':
        try:
            menu_item = Menu.objects.get(id=id)
        except Menu.DoesNotExist:
            return Response({'error': 'Menu item not found'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = MenuSerializer(menu_item)
        return Response({'menu': serializer.data}, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        try:
            menu_item = Menu.objects.get(id=id)
        except Menu.DoesNotExist:
            return Response({'error': 'Menu item not found'}, status=status.HTTP_404_NOT_FOUND)

        category = request.data.get('category')
        menu_name = request.data.get('name')
        price = request.data.get('price')
        
       
        if category:
      
            try:
                category_id = int(category) 
                category_instance = Categories.objects.get(id=category_id)  
            except ValueError:
                
                category_instance = Categories.objects.filter(name_category=category).first()
                if not category_instance:
                    return Response({'error': f'Category with name "{category}" not found.'}, status=status.HTTP_404_NOT_FOUND)
            except Categories.DoesNotExist:
                return Response({'error': 'Category not found'}, status=status.HTTP_404_NOT_FOUND)

            menu_item.category = category_instance

        if menu_name:
            menu_item.name = menu_name
        
        if price is not None:
            menu_item.price = price

        menu_item.save()

        serializer = MenuSerializer(menu_item)
        return Response({'message': 'Menu updated successfully!', 'menu': serializer.data}, status=status.HTTP_200_OK)

@api_view(['POST'])
def get_cycle_names(request):

    school_id = request.data.get('school_id')
    school_type = request.data.get('school_type')


    if not school_id or not school_type:
        return Response({'error': 'school_id and school_type are required.'}, status=status.HTTP_400_BAD_REQUEST)
    
    if school_type not in ['primary', 'secondary']:
        return Response({'error': 'Invalid school_type. It should be "primary" or "secondary".'}, status=status.HTTP_400_BAD_REQUEST)
    
    school_model = PrimarySchool if school_type == 'primary' else SecondarySchool
    school = school_model.objects.filter(id=school_id).first()

    if not school:
        return Response({'error': f'{school_type.capitalize()} School not found'}, status=status.HTTP_404_NOT_FOUND)

    if school_type == 'primary':
        menus = Menu.objects.filter(primary_school__id=school_id).values('cycle_name').distinct()
    elif school_type == 'secondary':
        menus = Menu.objects.filter(secondary_school__id=school_id).values('cycle_name').distinct()
    if not menus:
        return Response({'message': 'No cycle names found for this school.'}, status=status.HTTP_200_OK)

    cycle_names = [menu['cycle_name'] for menu in menus]

    return Response({
        'cycle_names': cycle_names
    }, status=status.HTTP_200_OK)



        
def get_custom_week_and_year():
    import datetime
    today = datetime.date.today()
    week_number = today.isocalendar()[1]  # Week number of the current year
    year = today.year
    return week_number, year



# MenuItems
@csrf_exempt
@api_view(['POST'])
@parser_classes([JSONParser])  # Use JSONParser here as you're sending JSON data
def add_menu_item(request):
    if request.method == 'POST':
        serializer = MenuItemsSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response({"error": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


@csrf_exempt
@api_view(['GET',])
def get_menu_items(request):
   if request.method == "GET":
        menu_items = MenuItems.objects.all()
        serializer = MenuItemsSerializer(menu_items,many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
   
@api_view(['GET', 'PUT', 'DELETE'])
def update_menu_items(request, pk):
    try:
        menu_item = MenuItems.objects.get(pk=pk)  
    except MenuItems.DoesNotExist:
        raise NotFound({'error': 'Menu item not found'}) 

    if request.method == 'GET':
    
        serializer = MenuItemsSerializer(menu_item)
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
      
        serializer = MenuItemsSerializer(menu_item, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()  
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
       
        menu_item.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# View Registered Students
@csrf_exempt
@api_view(['GET'])
def view_students(request):
    search_query = request.query_params.get('search', None)

    try:
        primary_students = PrimaryStudentsRegister.objects.all()
        secondary_students = SecondaryStudent.objects.all()

        if search_query:
            primary_students = primary_students.filter(
                student_name__icontains=search_query) | primary_students.filter(
                class_year__icontains=search_query)
            secondary_students = secondary_students.filter(
                secondary_student_name__icontains=search_query) | secondary_students.filter(
                secondary_class_year__icontains=search_query)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_404_NOT_FOUND)

    primary_students_serializer = PrimaryStudentSerializer(primary_students, many=True)
    primary_students_data = primary_students_serializer.data

    for student in primary_students_data:
        student['school_type'] = 'Primary'

        if student.get('parent'):
            parent = ParentRegisteration.objects.get(id=student['parent'])
            student['email'] = parent.email if parent else None
        else:
            staff = StaffRegisteration.objects.get(id=student['staff']) if student.get('staff') else None
            student['email'] = staff.email if staff else None

    secondary_students_serializer = SecondaryStudentSerializer(secondary_students, many=True)
    secondary_students_data = secondary_students_serializer.data

    for student in secondary_students_data:
        student['school_type'] = 'Secondary'

       
        if student.get('parent'):
            try:
                parent = ParentRegisteration.objects.get(id=student['parent'])
                student['email'] = parent.email if parent else student.get('email')
            except ParentRegisteration.DoesNotExist:
                pass
        elif student.get('staff'):
            try:
                staff = StaffRegisteration.objects.get(id=student['staff'])
                student['email'] = staff.email if staff else student.get('email')
            except StaffRegisteration.DoesNotExist:
                pass

    return Response({
        'primary_students': primary_students_data,
        'secondary_students': secondary_students_data
    }, status=status.HTTP_200_OK)




@api_view(['PUT', 'GET'])
def edit_student(request, student_id):
    try:
      
        student = PrimaryStudentsRegister.objects.get(id=student_id)
        school_type = 'primary'
        serializer = PrimaryStudentSerializer(student, data=request.data, partial=True)
    except PrimaryStudentsRegister.DoesNotExist:
     
        try:
            student = SecondaryStudent.objects.get(id=student_id)
            school_type = 'secondary'
            serializer = SecondaryStudentSerializer(student, data=request.data, partial=True)
        except SecondaryStudent.DoesNotExist:
            return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)
    
 
    if school_type not in ['primary', 'secondary']:
        return Response({'error': 'Invalid school type or student not found'}, status=status.HTTP_404_NOT_FOUND)
    
 
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@csrf_exempt
@api_view(['POST'])
def create_order(request):
    if request.method == 'POST':
        
        user_type = request.data.get('user_type')
        user_id = request.data.get('user_id')
        selected_days = request.data.get('selected_days')
        child_id = request.data.get('child_id', None)
        school_id = request.data.get('school_id', None)
        school_type = request.data.get('school_type', None)
        order_items_data = request.data.get('order_items', [])  

        # Validate the inputs
        if not user_type or not user_id or not selected_days:
            return Response({'error': 'User type, user ID, and selected days are required.'}, status=status.HTTP_400_BAD_REQUEST)

        if not order_items_data:
            return Response({'error': 'Order items are required.'}, status=status.HTTP_400_BAD_REQUEST)

        for item in order_items_data:
            if 'item_name' not in item or 'quantity' not in item or 'price' not in item:
                return Response({'error': 'Each order item must have item_name, quantity, and price.'}, status=status.HTTP_400_BAD_REQUEST)
        
    
        user = None
        if user_type == 'student':
            user = SecondaryStudent.objects.filter(id=user_id).first()
        elif user_type == 'parent':
            user = ParentRegisteration.objects.filter(id=user_id).first()
        elif user_type == 'staff':
            user = StaffRegisteration.objects.filter(id=user_id).first()

        if not user:
            return Response({'error': f'{user_type.capitalize()} not found.'}, status=status.HTTP_404_NOT_FOUND)
        
        if user_type == 'student' and not school_id:
            return Response({'error': 'School ID is required for students.'}, status=status.HTTP_400_BAD_REQUEST)

        if user_type == 'parent' and not school_id:
            return Response({'error': 'School ID is required for parents.'}, status=status.HTTP_400_BAD_REQUEST)

        created_orders = []  

        days_dict = {}
        for idx, day in enumerate(selected_days):
            if day not in days_dict:
                days_dict[day] = []
            if idx < len(order_items_data):
                days_dict[day].append(order_items_data[idx])

        for day, items_for_day in days_dict.items():
            menus_for_day = Menu.objects.filter(menu_day__iexact=day)

            if not menus_for_day:
                return Response({'error': f'No menus available for {day}.'}, status=status.HTTP_404_NOT_FOUND)

            order_total_price = 0
            order_items = []  # To store order items

            # Prepare the order data structure
            order_data = {
                'user_id': user.id,
                'user_type': user_type,
                'total_price': 0,
                'week_number': 0,  # To be calculated based on selected day
                'year': 0,
                'order_date': None,  # Set the actual order date
                'selected_day': day,
                'is_delivered': False,
                'status': 'pending',
            }

            if user_type in ['parent', 'staff'] and child_id:
                order_data['child_id'] = child_id
            if school_type == 'primary':
                order_data['primary_school'] = school_id
            elif school_type == 'secondary':
                order_data['secondary_school'] = school_id

            today = datetime.today()
            target_day = day.capitalize()  # Capitalizing to match the format in Menu model
            
            # Calculate the next occurrence of the target day
            target_day_num = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'].index(target_day)
            days_ahead = target_day_num - today.weekday()
            if days_ahead <= 0:  # If the target day is in the next week
                days_ahead += 7
            order_date = today + timedelta(days=days_ahead)

            # Calculate the week number of the target date
            week_number = order_date.isocalendar()[1]
            order_date = order_date.replace(hour=0, minute=0, second=0, microsecond=0)

            order_data['week_number'] = week_number
            order_data['year'] = order_date.year
            order_data['order_date'] = order_date

            # Create the order instance
            order_serializer = OrderSerializer(data=order_data)
            if not order_serializer.is_valid():
                return Response(order_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            order_instance = order_serializer.save()

            # Assign user_name based on user_type and child_id (for parent/staff)
            if user_type == 'parent' or user_type == 'staff':
                if child_id:
                    # Fetch the child's username if child_id is provided
                    child = PrimaryStudentsRegister.objects.filter(id=child_id).first()
                    if child:
                        order_instance.user_name = child.username  # Set child username
                else:
                    
                    order_instance.user_name = None  
            else:

                order_instance.user_name = user.username

            order_instance.save()

          
            for item in items_for_day:
                item_name = item['item_name']
                item_quantity = item['quantity']

                
                menu_item = menus_for_day.filter(name__iexact=item_name).first()

                if not menu_item:
                    return Response({'error': f'Menu item {item_name} not found for {day}.'}, status=status.HTTP_404_NOT_FOUND)

              
                order_item = OrderItem.objects.create(
                    menu=menu_item,  
                    quantity=item_quantity,  
                    order=order_instance
                )

                order_items.append(order_item) 
                order_total_price += menu_item.price * item_quantity  

            
            order_instance.total_price = order_total_price
            order_instance.save()

            
            order_details = {
                'order_id': order_instance.id,
                'selected_day': day,
                'total_price': order_instance.total_price,
                'order_date': order_instance.order_date.strftime('%d %b'), 
                'status': order_instance.status,
                'week_number': order_instance.week_number,
                'year': order_instance.year,
                'items': [
                    {
                        'item_name': order_item.menu.name,
                        'price': order_item.menu.price,
                        'quantity': order_item.quantity
                    } for order_item in order_items  
                ],
                'user_name': order_instance.user_name,  
            }

            if school_type == 'primary':
                order_details['school_id'] = school_id
                order_details['school_type'] = 'primary'
            elif school_type == 'secondary':
                order_details['school_id'] = school_id
                order_details['school_type'] = 'secondary'

        
            if child_id:
                order_details['child_id'] = child_id

            created_orders.append(order_details)  

        return Response({
            'message': 'Orders created successfully!',
            'orders': created_orders
        }, status=status.HTTP_201_CREATED)






@api_view(['POST'])
def complete_order(request):

    order_id = request.data.get('order_id')

    if not order_id:
        return Response({'error': 'Order ID is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        
        order = Order.objects.get(id=order_id)

        order.is_delivered = True
        order.status = 'collected'
        order.save()

        return Response({
            'message': 'Order completed successfully!',
            'order_id': order.id,
            'status': order.status,
            'is_delivered': order.is_delivered,
        }, status=status.HTTP_200_OK)

    except Order.DoesNotExist:
        return Response({'error': 'Order not found.'}, status=status.HTTP_404_NOT_FOUND)



@api_view(['POST'])
def cancel_order(request):
    order_id = request.data.get('order_id')

    if not order_id:
        return Response({'error': 'Order ID is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        order = Order.objects.get(id=order_id)

       
        order.status = 'cancelled'
        order.is_delivered = False
        order.save()

        credit_message = None
        user = None

        
        if order.user_type == 'student':
            user = SecondaryStudent.objects.get(id=order.user_id)
        elif order.user_type == 'parent':
            user = ParentRegisteration.objects.get(id=order.user_id)
        elif order.user_type == 'staff':
            user = StaffRegisteration.objects.get(id=order.user_id)

        if user:
            user.credits += order.total_price
            user.save()
            credit_message = f"Credits of {order.total_price} have been added to your account."

        order_info = {
            'order_id': order.id,
            'user_name': order.user_name,
            'total_price': order.total_price,
            'payment_method_id': getattr(order, 'payment_method_id', None),
            'status': order.status,
            'is_delivered': order.is_delivered,
            'credit_message': credit_message,
            'user_credits': user.credits if user else None,
        }

        return Response({
            'message': 'Order cancelled successfully!',
            'order_info': order_info
        }, status=status.HTTP_200_OK)

    except Order.DoesNotExist:
        return Response({'error': 'Order not found.'}, status=status.HTTP_404_NOT_FOUND)
    except (ParentRegisteration.DoesNotExist, StaffRegisteration.DoesNotExist, SecondaryStudent.DoesNotExist):
        return Response({'error': 'User not found for credits update.'}, status=status.HTTP_404_NOT_FOUND)



@csrf_exempt
@api_view(['POST'])
def add_order_item(request):
    if request.method=='POST':
        serializer=OrderItemSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data,status=status.HTTP_201_CREATED, )
        else:
            return Response({ "error" : serializer.errors},status = status.HTTP_400_BAD_REQUEST)
        
def safe_localtime(dt):
    """Ensure datetime is timezone-aware before converting to localtime."""
    if isinstance(dt, date) and not isinstance(dt, datetime):
        dt = make_aware(datetime.combine(dt, datetime.min.time()))
    return localtime(dt)

@api_view(['GET'])
def get_all_orders(request):
    order_details = []

    # Regular orders
    orders = Order.objects.all().order_by('-order_date')
    for order in orders:
        order_items = OrderItem.objects.filter(order=order)
        items_details = []
        for item in order_items:
            item_name = item._menu_name if item._menu_name else (item.menu.name if item.menu else "Deleted Menu")
            item_price = item._menu_price if item._menu_price else (item.menu.price if item.menu else 0)
            items_details.append({
                'item_name': item_name,
                'item_price': item_price,
                'quantity': item.quantity
            })

        local_order_date = safe_localtime(order.order_date)
        formatted_order_date = local_order_date.strftime('%d %B, %Y')

        order_data = {
            'order_id': order.id,
            'selected_day': order.selected_day,
            'total_price': order.total_price,
            'order_date': formatted_order_date,
            'status': order.status,
            'week_number': order.week_number,
            'year': order.year,
            'items': items_details,
            'user_name': order.user_name,
            'user_type': order.user_type,
            'payment_id': order.payment_id,
        }

        # Add school info
        if order.primary_school:
            order_data['school_id'] = order.primary_school.id
            order_data['school_name'] = order.primary_school.school_name
            order_data['school_type'] = 'primary'
        elif order.secondary_school:
            order_data['school_id'] = order.secondary_school.id
            order_data['school_name'] = order.secondary_school.secondary_school_name
            order_data['school_type'] = 'secondary'

        if order.user_type in ['parent', 'staff']:
            order_data['child_id'] = order.child_id

        order_details.append(order_data)

    # Manager orders
    manager_orders = ManagerOrder.objects.all().order_by('-order_date')
    for m_order in manager_orders:
        m_items = ManagerOrderItem.objects.filter(order=m_order)
        items_details = [
            {
                'day': item.day,
                'item': item.item,
                'quantity': item.quantity,
                'production_price': item.production_price or 0
            }
            for item in m_items
        ]

        local_order_date = safe_localtime(m_order.order_date)
        formatted_order_date = local_order_date.strftime('%d %B, %Y')

        order_data = {
            'order_id': f"m_{m_order.id}",
            'selected_day': m_order.selected_day,
            'total_production_price': m_order.total_production_price,
            'order_date': formatted_order_date,
            'status': m_order.status,
            'week_number': m_order.week_number,
            'year': m_order.year,
            'is_delivered': m_order.is_delivered,
            'user_name': m_order.manager.username if hasattr(m_order.manager, 'username') else str(m_order.manager),
            'user_type': 'manager',
            'items': items_details,
        }

        # Add school info for manager orders if they have a school
        if hasattr(m_order.manager, 'primary_school') and m_order.manager.primary_school:
            order_data['school_id'] = m_order.manager.primary_school.id
            order_data['school_name'] = m_order.manager.primary_school.school_name
            order_data['school_type'] = 'primary'
        elif hasattr(m_order.manager, 'secondary_school') and m_order.manager.secondary_school:
            order_data['school_id'] = m_order.manager.secondary_school.id
            order_data['school_name'] = m_order.manager.secondary_school.secondary_school_name
            order_data['school_type'] = 'secondary'

        order_details.append(order_data)

    return Response({
        'message': 'Orders retrieved successfully!',
        'orders': order_details
    }, status=status.HTTP_200_OK)




@api_view(['GET'])
def get_order_by_id(request, order_id):
    if str(order_id).startswith('m_'):
        manager_order_id = str(order_id).replace('m_', '')
        try:
            order = ManagerOrder.objects.get(id=manager_order_id)
            order_items = ManagerOrderItem.objects.filter(order=order)

            items_details = [
                {
                    'day': item.day,
                    'item': item.item,
                    'quantity': item.quantity,
                    'production_price': item.production_price or 0
                }
                for item in order_items
            ]

            local_order_date = safe_localtime(order.order_date)
            formatted_order_date = local_order_date.strftime('%d %B, %Y')

            order_data = {
                'order_id': f"m_{order.id}",
                'selected_day': order.selected_day,
                'total_production_price': order.total_production_price,
                'order_date': formatted_order_date,
                'status': order.status,
                'week_number': order.week_number,
                'year': order.year,
                'is_delivered': order.is_delivered,
                'manager_name': order.manager.username if hasattr(order.manager, 'username') else str(order.manager),
                'items': items_details,
                'user_type': 'manager'
            }

            return Response(
                {'message': 'Manager order retrieved successfully!', 'order': order_data},
                status=status.HTTP_200_OK
            )

        except ManagerOrder.DoesNotExist:
            return Response({'error': 'Manager order not found.'}, status=status.HTTP_404_NOT_FOUND)

    # ---------------- NORMAL ORDER ----------------
    try:
        order = Order.objects.get(id=order_id)
        order_items = OrderItem.objects.filter(order=order)

        items_details = []
        for item in order_items:
            item_name = item._menu_name if item._menu_name else (item.menu.name if item.menu else "Deleted Menu")
            item_price = item._menu_price if item._menu_price else (item.menu.price if item.menu else 0)
            items_details.append({
                'item_name': item_name,
                'item_price': item_price,
                'quantity': item.quantity
            })

        local_order_date = safe_localtime(order.order_date)
        formatted_order_date = local_order_date.strftime('%d %B, %Y')

        order_data = {
            'order_id': order.id,
            'selected_day': order.selected_day,
            'total_price': order.total_price,
            'order_date': formatted_order_date,
            'status': order.status,
            'week_number': order.week_number,
            'year': order.year,
            'items': items_details,
            'user_name': order.user_name,
        }

        if order.user_type in ['parent', 'staff']:
            order_data['child_id'] = order.child_id

        if order.primary_school:
            order_data['school_id'] = order.primary_school.id
            order_data['school_type'] = 'primary'
        elif order.secondary_school:
            order_data['school_id'] = order.secondary_school.id
            order_data['school_type'] = 'secondary'

        return Response({'message': 'Order retrieved successfully!', 'order': order_data},
                        status=status.HTTP_200_OK)

    except Order.DoesNotExist:
        return Response({'error': 'Order not found.'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
def get_orders_by_school(request):
    school_id = request.data.get('school_id')
    school_type = request.data.get('school_type')

    if not school_id or not school_type:
        return Response({'error': 'Both school_id and school_type are required.'}, status=status.HTTP_400_BAD_REQUEST)

    if school_type not in ['primary', 'secondary']:
        return Response({'error': 'Invalid school_type. It should be either "primary" or "secondary".'},
                        status=status.HTTP_400_BAD_REQUEST)

    if school_type == 'primary':
        orders = Order.objects.filter(primary_school_id=school_id).order_by('-order_date')
    else:
        orders = Order.objects.filter(secondary_school_id=school_id).order_by('-order_date')

    if school_type == 'primary':
        manager_orders = ManagerOrder.objects.filter(
            manager__school_type='primary',
            manager__primary_school_id=school_id
        ).order_by('-order_date')
    else:
        manager_orders = ManagerOrder.objects.filter(
            manager__school_type='secondary',
            manager__secondary_school_id=school_id
        ).order_by('-order_date')


    if not orders.exists() and not manager_orders.exists():
        return Response({'error': 'No orders found for the given school.'}, status=status.HTTP_404_NOT_FOUND)

    order_details = []

    # Normal Orders
    for order in orders:
        order_items = order.order_items.all()
        items_details = []
        for item in order_items:
            item_name = item._menu_name if item._menu_name else (item.menu.name if item.menu else "Deleted Menu")
            item_price = item._menu_price if item._menu_price else (item.menu.price if item.menu else 0)
            items_details.append({
                'item_name': item_name,
                'item_price': item_price,
                'quantity': item.quantity
            })

        local_order_date = safe_localtime(order.order_date)
        formatted_order_date = local_order_date.strftime('%d %B, %Y')

        order_data = {
            'order_id': order.id,
            'selected_day': order.selected_day,
            'total_price': order.total_price,
            'order_date': formatted_order_date,
            'status': order.status,
            'week_number': order.week_number,
            'year': order.year,
            'items': items_details,
            'user_name': order.user_name,
            'user_type': order.user_type,
            'school_id': school_id,
            'school_type': school_type
        }

        if order.user_type in ['parent', 'staff']:
            order_data['child_id'] = order.child_id

        order_details.append(order_data)

    # Manager Orders
    for m_order in manager_orders:
        order_items = ManagerOrderItem.objects.filter(order=m_order)
        items_details = [
            {
                'day': item.day,
                'item': item.item,
                'quantity': item.quantity,
                'production_price': item.production_price or 0
            }
            for item in order_items
        ]

        local_order_date = safe_localtime(m_order.order_date)
        formatted_order_date = local_order_date.strftime('%d %B, %Y')

        order_data = {
            'order_id': f"m_{m_order.id}",
            'selected_day': m_order.selected_day,
            'total_production_price': m_order.total_production_price,
            'order_date': formatted_order_date,
            'status': m_order.status,
            'week_number': m_order.week_number,
            'year': m_order.year,
            'is_delivered': m_order.is_delivered,
            'manager_name': m_order.manager.username if hasattr(m_order.manager, 'username') else str(m_order.manager),
            'items': items_details,
            'user_type': 'manager',
            'school_id': school_id,
            'school_type': school_type
        }

        order_details.append(order_data)

    return Response({'message': 'Orders retrieved successfully!', 'orders': order_details},
                    status=status.HTTP_200_OK)



@api_view(['POST'])
def get_orders_by_user(request):
    user_id = request.data.get('user_id')
    user_type = request.data.get('user_type')
    child_id = request.data.get('child_id')

    if not user_type:
        return Response({'error': 'user_type is required.'}, status=status.HTTP_400_BAD_REQUEST)

    if user_type not in ['student', 'parent', 'staff', 'manager']:
        return Response({'error': 'Invalid user_type.'}, status=status.HTTP_400_BAD_REQUEST)

    order_details = []
    orders = Order.objects.none()

    # Normal user logic
    if user_type == 'staff' and child_id:
        orders = Order.objects.filter(child_id=child_id).order_by('-order_date')
    elif user_type == 'staff' and user_id:
        orders = Order.objects.filter(user_id=user_id, user_type='staff').order_by('-order_date')
    elif user_type == 'parent':
        orders = Order.objects.filter(user_id=user_id, user_type='parent').order_by('-order_date')
    elif user_type == 'student':
        orders = Order.objects.filter(user_id=user_id, user_type='student').order_by('-order_date')

    for order in orders:
        order_items = order.order_items.all()
        items_details = []
        for item in order_items:
            item_name = item._menu_name if item._menu_name else (item.menu.name if item.menu else "Deleted Menu")
            item_price = item._menu_price if item._menu_price else (item.menu.price if item.menu else 0)
            items_details.append({
                'item_name': item_name,
                'item_price': item_price,
                'quantity': item.quantity
            })

        local_date = safe_localtime(order.order_date).date()
        formatted_order_date = local_date.strftime('%d %B, %Y')

        order_data = {
            'order_id': order.id,
            'selected_day': order.selected_day,
            'total_price': order.total_price,
            'order_date': formatted_order_date,
            'order_date_raw': str(local_date),
            'status': order.status,
            'week_number': order.week_number,
            'year': order.year,
            'items': items_details,
            'user_name': order.user_name,
            'user_type': order.user_type,
        }

        if order.user_type in ['parent', 'staff']:
            order_data['child_id'] = order.child_id

        if order.primary_school:
            order_data['school_id'] = order.primary_school.id
            order_data['school_type'] = 'primary'
        elif order.secondary_school:
            order_data['school_id'] = order.secondary_school.id
            order_data['school_type'] = 'secondary'

        order_details.append(order_data)

    # Manager Orders
    if user_type == 'manager' and user_id:
        manager_orders = ManagerOrder.objects.filter(manager_id=user_id).order_by('-order_date')

        for m_order in manager_orders:
            order_items = ManagerOrderItem.objects.filter(order=m_order)
            items_details = [
                {
                    'day': item.day,
                    'item': item.item,
                    'quantity': item.quantity,
                    'production_price': item.production_price or 0
                }
                for item in order_items
            ]

            local_order_date = safe_localtime(m_order.order_date)
            formatted_order_date = local_order_date.strftime('%d %B, %Y')

            order_data = {
                'order_id': f"m_{m_order.id}",
                'selected_day': m_order.selected_day,
                'total_production_price': m_order.total_production_price,
                'order_date': formatted_order_date,
                'order_date_raw': str(local_order_date.date()),
                'status': m_order.status,
                'week_number': m_order.week_number,
                'year': m_order.year,
                'is_delivered': m_order.is_delivered,
                'manager_name': m_order.manager.username,
                'items': items_details,
                'user_type': 'manager',
            }

            if m_order.manager.school_type == 'primary' and m_order.manager.primary_school:
                order_data['school_id'] = m_order.manager.primary_school.id
                order_data['school_type'] = 'primary'
            elif m_order.manager.school_type == 'secondary' and m_order.manager.secondary_school:
                order_data['school_id'] = m_order.manager.secondary_school.id
                order_data['school_type'] = 'secondary'

            order_details.append(order_data)

    if not order_details:
        return Response({'error': 'No orders found for the given user.'}, status=status.HTTP_404_NOT_FOUND)

    return Response({'message': 'Orders retrieved successfully!', 'orders': order_details},
                    status=status.HTTP_200_OK)


@api_view(['POST'])
def contactmessage(request):
    """
    Handles the contact form submission. It validates the form, saves the data,
    sends an email notification, and returns a success or error response.
    """
    full_name = request.data.get('full_name')
    email = request.data.get('email')
    phone = request.data.get('phone')
    subject = request.data.get('subject')
    message = request.data.get('message')
    photo = request.FILES.get('photo')  
    
    if not full_name or not email or not subject or not message:
        return Response({"error": "Full name, email, subject, and message are required."}, status=status.HTTP_400_BAD_REQUEST)

    if not re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
        return Response({"error": "Invalid email format."}, status=status.HTTP_400_BAD_REQUEST)

    if phone and not re.match(r'^\+?[0-9\s\-]+$', phone):
        return Response({"error": "Invalid phone number format."}, status=status.HTTP_400_BAD_REQUEST)

    photo_filename = None
    if photo:
        file_extension = photo.name.split('.')[-1].lower()
        if file_extension not in ALLOWED_FILE_EXTENSIONS:
            return Response({"error": "Invalid file type. Only PNG, JPG, JPEG, and GIF files are allowed."}, status=status.HTTP_400_BAD_REQUEST)

        if photo.size > 1024 * 1024:
            return Response({"error": "File size too large. Maximum size is 1MB."}, status=status.HTTP_400_BAD_REQUEST)

        fs = FileSystemStorage(location='media/contact_photos')
        photo_filename = fs.save(photo.name, photo)

    try:
        contact_message = ContactMessage.objects.create(
            full_name=full_name,
            email=email,
            phone=phone,
            subject=subject,
            message=message,
            photo_filename=photo_filename if photo_filename else ""  
        )

        # HTML email content
        html_message = f"""
        <html>
            <body>
                <h2 style="color:#4CAF50;">New Contact Message</h2>
                <p><strong>Name:</strong> {full_name}</p>
                <p><strong>Email:</strong> {email}</p>
                <p><strong>Phone:</strong> {phone if phone else 'Not provided'}</p>
                <p><strong>Subject:</strong> {subject}</p>
                <p><strong>Message:</strong></p>
                <blockquote style="border-left: 4px solid #4CAF50; padding-left: 15px;">
                    {message}
                </blockquote>
                <p><em>We will respond to you shortly. Thank you for reaching out!</em></p>
            </body>
        </html>
        """

        send_mail(
            subject=f"New Contact Message: {subject}",
            message=message,  # Plain text message (you can remove this if you want only HTML email)
            from_email=settings.DEFAULT_FROM_EMAIL,  
            recipient_list=[settings.MAIL_DEFAULT_SENDER],  # This will send to the email set in MAIL_DEFAULT_SENDER
            fail_silently=False,
            html_message=html_message  # Include HTML content
        )

        return Response({"message": "Thank you for your message! We will get back to you shortly."}, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response({"error": f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def top_up_credits(request):
    """
    Top-up credits for the parent, staff, or student.
    """
    serializer = TopUpCreditsSerializer(data=request.data)
    
    if serializer.is_valid():
        user_id = serializer.validated_data['user_id']
        amount = serializer.validated_data['amount']
        user_type = serializer.validated_data['user_type']

      
        if user_type == "parent":
            user = ParentRegisteration.objects.filter(id=user_id).first()
        elif user_type == "staff":
            user = StaffRegisteration.objects.filter(id=user_id).first()
        elif user_type == "student":
            user = SecondaryStudent.objects.filter(id=user_id).first()

        if not user:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)

    
        user.top_up_credits(amount)

        return Response({"message": f"{user_type.capitalize()} credits successfully updated."}, status=status.HTTP_200_OK)

    return Response({"error": "Invalid data"}, status=status.HTTP_400_BAD_REQUEST)



def get_current_week_and_year():
    """
    Get the current week number and year.
    """
    current_date = datetime.now()
    current_week = current_date.isocalendar()[1]  
    current_year = current_date.year
    return current_week, current_year


def get_or_create_customer(email, name):
    """Find or create a Stripe customer"""
    customers = stripe.Customer.list(email=email).data
    if customers:
        return customers[0]  # return existing
    return stripe.Customer.create(email=email, name=name)

class CreateOrderAndPaymentAPIView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            data = request.data
            user_type = data.get('user_type')
            user_id = data.get('user_id')
            selected_days = data.get('selected_days')
            order_items_data = data.get('order_items', [])
            school_id = data.get('school_id', None)
            school_type = data.get('school_type', None)
            child_id = data.get('child_id', None)
            payment_id = data.get("payment_id", None)
            front_end_total_price = float(data.get("total_price", 0))

            if not user_type or not user_id or not selected_days:
                return Response({'error': 'User type, user ID, and selected days are required.'},
                                status=status.HTTP_400_BAD_REQUEST)

            if not order_items_data:
                return Response({'error': 'Order items are required.'}, status=status.HTTP_400_BAD_REQUEST)

            # Ensure list lengths match
            if len(selected_days) != len(order_items_data):
                return Response({'error': 'The number of selected days must match the number of order items.'},
                                status=status.HTTP_400_BAD_REQUEST)

            # Get user instance
            if user_type == 'student':
                user = SecondaryStudent.objects.filter(id=user_id).first()
            elif user_type == 'parent':
                user = ParentRegisteration.objects.filter(id=user_id).first()
            elif user_type == 'staff':
                user = StaffRegisteration.objects.filter(id=user_id).first()
            else:
                return Response({'error': 'Invalid user type.'}, status=status.HTTP_400_BAD_REQUEST)

            if not user:
                return Response({'error': f'{user_type.capitalize()} not found.'}, status=status.HTTP_404_NOT_FOUND)

            if user_type in ['student', 'parent'] and not school_id:
                return Response({'error': 'School ID is required for students and parents.'},
                                status=status.HTTP_400_BAD_REQUEST)

            # Primary students/parents get free meals → all prices forced to 0
            is_primary_free = school_type == "primary"

            # Transaction for atomicity
            with transaction.atomic():
                created_orders = []
                calculated_total_price = 0

                # Group items by day
                daily_orders = {}
                for i, day in enumerate(selected_days):
                    item_for_day = order_items_data[i]
                    daily_orders.setdefault(day, []).append(item_for_day)

                # Create orders
                for day, items_list in daily_orders.items():
                    today = datetime.now()
                    target_day_num = ['monday', 'tuesday', 'wednesday', 'thursday',
                                    'friday', 'saturday', 'sunday'].index(day.lower())
                    days_ahead = (target_day_num - today.weekday() + 7) % 7
                    if today.weekday() >= 4: 
                        days_ahead += 7

                    order_date = today + timedelta(days=days_ahead)

                    week_number = order_date.isocalendar()[1]
                    order_date = order_date.replace(hour=0, minute=0, second=0, microsecond=0)

                    order_instance = Order.objects.create(
                        user_id=user.id,
                        user_type=user_type,
                        total_price=0,
                        week_number=week_number,
                        year=order_date.year,
                        order_date=order_date,
                        selected_day=day,
                        is_delivered=False,
                        status='pending',
                        payment_id=payment_id,
                        child_id=child_id if user_type in ['parent', 'staff'] and child_id else None,
                        primary_school_id=school_id if school_type == 'primary' else None,
                        secondary_school_id=school_id if school_type == 'secondary' else None
                    )

                    daily_total_price = 0

                    # Create OrderItems
                    for item in items_list:
                        item_name = item.get('item_name')
                        quantity = item.get('quantity')

                        if not item_name or not quantity:
                            return Response({'error': 'Each item must have item_name and quantity.'},
                                            status=status.HTTP_400_BAD_REQUEST)

                        menu_item = Menu.objects.filter(menu_day__iexact=day, name__iexact=item_name).first()
                        if not menu_item:
                            return Response({'error': f'Menu item with name "{item_name}" not found for {day}.'},
                                            status=status.HTTP_404_NOT_FOUND)

                    
                        item_price = 0 if is_primary_free else float(item.get('price', menu_item.price))

                        OrderItem.objects.create(
                            order=order_instance,
                            quantity=quantity,
                            _menu_name=menu_item.name,
                            _menu_price=item_price
                        )

                        daily_total_price += item_price * quantity

                    order_instance.total_price = daily_total_price
                    order_instance.save()

                    calculated_total_price += daily_total_price
                    created_orders.append(order_instance)

   
                if is_primary_free:
                    calculated_total_price = 0


                # ---------------------------
                # Handle primary (free meals)
                # ---------------------------
                if is_primary_free:
                    for order in created_orders:
                        order.status = 'pending'
                        order.save()

                    return Response({
                        'message': 'Orders created successfully with free meal for child.',
                        'orders': OrderSerializer(created_orders, many=True).data
                    }, status=status.HTTP_201_CREATED)

                # ---------------------------
                # Handle credits (secondary without payment_id)
                # ---------------------------
                if not payment_id:
                    if user.credits < calculated_total_price:
                        raise ValueError("Insufficient credits to complete the order.")

                    user.credits -= calculated_total_price
                    user.save()

                    for order in created_orders:
                        order.status = 'pending'
                        order.save()

                    return Response({
                        'message': 'Orders created and credits deducted successfully!',
                        'orders': OrderSerializer(created_orders, many=True).data
                    }, status=status.HTTP_201_CREATED)

                # ---------------------------
                # Handle Stripe payment (secondary with payment_id)
                # ---------------------------
                total_price_in_cents = int(calculated_total_price * 100)

                customers = stripe.Customer.list(email=user.email).data
                if customers:
                    customer = customers[0]
                else:
                    customer = stripe.Customer.create(
                        email=user.email,
                        name=f"{getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')}".strip()
                    )

                stripe.PaymentMethod.attach(payment_id, customer=customer.id)
                stripe.Customer.modify(customer.id, invoice_settings={"default_payment_method": payment_id})

                payment_intent = stripe.PaymentIntent.create(
                    amount=total_price_in_cents,
                    currency="eur",
                    customer=customer.id,
                    payment_method=payment_id,
                    confirmation_method="manual",
                    confirm=True,
                    receipt_email=user.email,
                    return_url=f"{request.scheme}://{request.get_host()}/payment-success/",
                )

                for order in created_orders:
                    order.payment_id = payment_intent.id
                    order.status = 'pending'
                    order.save()

                return Response({
                    'message': 'Orders and payment intent created successfully!',
                    'orders': OrderSerializer(created_orders, many=True).data,
                    'payment_intent': payment_intent.client_secret
                }, status=status.HTTP_201_CREATED)

        except stripe.error.CardError as e:
            return Response({"error": f"Card Error: {e.user_message}"}, status=status.HTTP_400_BAD_REQUEST)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




@api_view(['POST'])
def top_up_payment(request):
    """
    API to process top-up credits for a parent, staff, or student.
    """
    try:
        user_id = request.data.get('user_id')
        amount = request.data.get('amount')
        user_type = request.data.get('user_type')
        payment_method_id = request.data.get('payment_method_id')

        if not all([user_id, amount, user_type, payment_method_id]):
            return Response({"error": "Missing required parameters."}, status=status.HTTP_400_BAD_REQUEST)

        # ✅ Get user
        if user_type == "parent":
            user = ParentRegisteration.objects.filter(id=user_id).first()
        elif user_type == "staff":
            user = StaffRegisteration.objects.filter(id=user_id).first()
        elif user_type == "student":
            user = SecondaryStudent.objects.filter(id=user_id).first()
        else:
            return Response({"error": "Invalid user type."}, status=status.HTTP_400_BAD_REQUEST)

        if not user:
            return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)

        # ✅ Amount in cents
        try:
            amount_in_cents = int(float(amount) * 100)
        except ValueError:
            return Response({"error": "Invalid amount."}, status=status.HTTP_400_BAD_REQUEST)

        # ✅ Ensure customer exists
        if not getattr(user, "stripe_customer_id", None):
            customer = stripe.Customer.create(
                name=f"{user.first_name} {user.last_name}",
                email=user.email,
            )
            user.stripe_customer_id = customer.id
            user.save()

        # ✅ Attach payment method
        stripe.PaymentMethod.attach(
            payment_method_id,
            customer=user.stripe_customer_id,
        )

        # ✅ Update default payment method
        stripe.Customer.modify(
            user.stripe_customer_id,
            invoice_settings={"default_payment_method": payment_method_id}
        )

        # ✅ Create PaymentIntent (card only, no redirects)
        payment_intent = stripe.PaymentIntent.create(
            amount=amount_in_cents,
            currency="eur",
            customer=user.stripe_customer_id,
            payment_method=payment_method_id,
            payment_method_types=["card"],   # 👈 restrict to card only
            confirmation_method="manual",
            confirm=True,
            receipt_email=user.email,        # 📩 receipt will have name + email
        )

        # ✅ On success → top up credits
        if payment_intent.status == 'succeeded':
            user.top_up_credits(float(amount))
            return Response({
                "message": f"{user_type.capitalize()} credits successfully updated.",
                "credits": user.credits,
                "payment_id": payment_intent.id
            }, status=status.HTTP_200_OK)
        else:
            return Response({"error": "Payment failed."}, status=status.HTTP_400_BAD_REQUEST)

    except stripe.error.CardError as e:
        return Response({"error": f"Card Error: {e.user_message}"}, status=status.HTTP_400_BAD_REQUEST)
    except stripe.error.StripeError as e:
        return Response({"error": f"Stripe error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        return Response({"error": f"Error processing payment: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def get_custom_week_and_year():
  
    today = datetime.today()
    return today.isocalendar()[1], today.year  

DAY_COLORS = {
    "Monday": "FF0000",
    "Tuesday": "00FF00",
    "Wednesday": "0000FF",
    "Thursday": "FFFF00",
    "Friday": "FF00FF",
}

CLASS_YEARS = ["1st", "2nd", "3rd", "4th", "5th", "6th"]

def fetch_orders(school_id, school_type, target_day=None):
    current_time = datetime.now()
    current_week_number = current_time.isocalendar()[1]
    current_year = current_time.year
    
    cutoff_day = 4  
    cutoff_hour = 2 

    is_past_cutoff = (
        current_time.weekday() > cutoff_day or
        (current_time.weekday() == cutoff_day and current_time.hour >= cutoff_hour)
    )

    target_week = current_week_number + 1 if is_past_cutoff else current_week_number
    target_year = current_year
    
    if target_week > 52:
        target_week = 1
        target_year += 1
    
    filter_kwargs = {"week_number": target_week, "year": target_year}
    
    if school_type == "primary":
        filter_kwargs["primary_school_id"] = school_id
    else:
        filter_kwargs["secondary_school_id"] = school_id

    if target_day:
        filter_kwargs["selected_day__iexact"] = target_day

    orders = (
        Order.objects
        .filter(**filter_kwargs)
        .exclude(status__iexact="cancelled") 
        .order_by("selected_day")
    )
    
    student_orders = orders.exclude(user_type="staff")
    staff_orders = orders.filter(user_type="staff")

    return student_orders, staff_orders

def generate_workbook(school, student_orders, staff_orders, school_type, role='admin', day_filter=None):
    workbook = Workbook()
    workbook.remove(workbook.active)

    # === Data Stores ===
    day_totals = defaultdict(lambda: defaultdict(int))
    grouped_orders = defaultdict(lambda: defaultdict(list))
    staff_orders_by_day = defaultdict(list)
    teacher_totals = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    class_totals = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    # === Styles ===
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    title_font = Font(bold=True, size=14, color="000000")
    title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    def apply_header_styling(sheet, title_text, columns):
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(columns))
        title_cell = sheet.cell(row=1, column=1, value=title_text)
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center_align

        for col_num, column_title in enumerate(columns, 1):
            cell = sheet.cell(row=2, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
            sheet.column_dimensions[get_column_letter(col_num)].width = 25

    def apply_data_styling(sheet, start_row):
        for row in sheet.iter_rows(min_row=start_row):
            for cell in row:
                cell.border = border
                cell.alignment = left_align if cell.column == 1 else center_align

    # === Process Student Orders ===
    for order in student_orders:
        selected_day = order.selected_day
        order_items = order.order_items.all()

        item_data = {
            (item._menu_name or (item.menu.name if item.menu else "Deleted Menu")): item.quantity
            for item in order_items
        }

        order_data = {
            "order_id": order.id,
            "student_name": "Unknown",
            "class_year": None,
            "teacher_name": None,
            "order_items": item_data,
        }

        if school_type == "primary":
            student = PrimaryStudentsRegister.objects.filter(id=order.child_id).first()
            if student:
                full_name = f"{(student.first_name or '').strip()} {(student.last_name or '').strip()}".strip()
                order_data["student_name"] = full_name or (student.email.split("@")[0] if student.email else "Unknown")
                teacher_name = student.teacher.teacher_name if student.teacher else "Unknown"
                order_data["teacher_name"] = teacher_name
                grouped_orders[selected_day][teacher_name].append(order_data)
                for menu_name, quantity in item_data.items():
                    teacher_totals[selected_day][teacher_name][menu_name] += quantity
                    day_totals[selected_day][menu_name] += quantity
        else:
            student = SecondaryStudent.objects.filter(id=order.user_id).first()
            if student:
                full_name = f"{(student.first_name or '').strip()} {(student.last_name or '').strip()}".strip()
                order_data["student_name"] = full_name or (student.email.split("@")[0] if student.email else "Unknown")
                class_year = student.class_year if student.class_year else "Unknown"
                order_data["class_year"] = class_year
                grouped_orders[selected_day][class_year].append(order_data)
                for menu_name, quantity in item_data.items():
                    class_totals[selected_day][class_year][menu_name] += quantity
                    day_totals[selected_day][menu_name] += quantity

    # === Process Staff Orders ===
    for order in staff_orders:
        selected_day = order.selected_day
        order_items = order.order_items.all()
        item_data = {
            (item._menu_name or (item.menu.name if item.menu else "Deleted Menu")): item.quantity
            for item in order_items
        }
        staff_order_data = {"order_id": order.id, "staff_name": "Unknown", "order_items": item_data}
        staff = StaffRegisteration.objects.filter(id=order.user_id).first()
        if staff:
            staff_order_data["staff_name"] = staff.username
        staff_orders_by_day[selected_day].append(staff_order_data)
        for menu_name, quantity in item_data.items():
            day_totals[selected_day][menu_name] += quantity

    # === Generate Sheets ===
    all_days = list(DAY_COLORS.keys())
    days_to_generate = [day_filter] if day_filter in all_days else all_days
    has_visible_sheet = False

    for day in days_to_generate:
        # === Class/Teacher Sheets (Admin + Staff) ===
        if role in ["admin", "staff"]:
            entity_list = Teacher.objects.filter(school=school) if school_type == "primary" else CLASS_YEARS
            for entity in entity_list:
                entity_name = entity.teacher_name if school_type == "primary" else entity
                sheet = workbook.create_sheet(title=f"{entity_name} - {day}"[:31])
                sheet.sheet_properties.tabColor = DAY_COLORS.get(day, "FFFFFF")
                has_visible_sheet = True

                title = (
                    f"{entity_name} Order Sheet for {day} of {school}"
                    if school_type == "primary"
                    else f"Class {entity_name} Order Sheet for {day} of {school}"
                )
                apply_header_styling(sheet, title, ["Order ID", "Student Name", "Menu Items", "Quantity"])

                row_num = 3
                for order_data in grouped_orders.get(day, {}).get(entity_name, []):
                    for menu_name, quantity in order_data["order_items"].items():
                        sheet.cell(row=row_num, column=1, value=order_data["order_id"])
                        sheet.cell(row=row_num, column=2, value=order_data["student_name"])
                        sheet.cell(row=row_num, column=3, value=menu_name)
                        sheet.cell(row=row_num, column=4, value=quantity)
                        row_num += 1
                if row_num == 3:
                    sheet.cell(row=3, column=1, value="No orders")
                    sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=4)
                apply_data_styling(sheet, 3)

        # === Teacher/Class Totals ===
        if role in ["admin", "chef", "staff"]:
            if school_type == "primary":
                totals = teacher_totals
                label = "Teacher"
            else:
                totals = class_totals
                label = "Class"

            sheet = workbook.create_sheet(title=f"{day} {label} Totals"[:31])
            sheet.sheet_properties.tabColor = "00B0F0"
            apply_header_styling(sheet, f"{label} Totals for {day} ({school})", [f"{label} Name", "Menu Item", "Total Quantity"])
            row_num = 3
            day_grand_total = 0

            if totals.get(day):
                for name, items in totals[day].items():
                    subtotal = 0
                    for menu_name, qty in items.items():
                        sheet.cell(row=row_num, column=1, value=name)
                        sheet.cell(row=row_num, column=2, value=menu_name)
                        sheet.cell(row=row_num, column=3, value=qty)
                        subtotal += qty
                        day_grand_total += qty
                        row_num += 1
                    sheet.cell(row=row_num, column=1, value=f"Total for {name}")
                    sheet.merge_cells(start_row=row_num, end_row=row_num, start_column=1, end_column=2)
                    total_cell = sheet.cell(row=row_num, column=3, value=subtotal)
                    total_cell.font = Font(bold=True)
                    total_cell.fill = total_fill
                    row_num += 1

                row_num += 1
                sheet.cell(row=row_num, column=1, value="Grand Total (All)")
                sheet.merge_cells(start_row=row_num, end_row=row_num, start_column=1, end_column=2)
                cell = sheet.cell(row=row_num, column=3, value=day_grand_total)
                cell.font = Font(bold=True)
                cell.fill = total_fill
            else:
                sheet.cell(row=3, column=1, value="No data available")
                sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=3)
            apply_data_styling(sheet, 3)

        # === Staff Sheet (Admin Only) ===
        if role == "admin":
            sheet = workbook.create_sheet(title=f"Staff {day}"[:31])
            sheet.sheet_properties.tabColor = "CCCCCC"
            apply_header_styling(sheet, f"Staff Order Sheet for {day} of {school}", ["Order ID", "Staff Name", "Menu Item", "Quantity"])
            row_num = 3
            if staff_orders_by_day.get(day):
                for order_data in staff_orders_by_day[day]:
                    for menu_name, quantity in order_data["order_items"].items():
                        sheet.cell(row=row_num, column=1, value=order_data["order_id"])
                        sheet.cell(row=row_num, column=2, value=order_data["staff_name"])
                        sheet.cell(row=row_num, column=3, value=menu_name)
                        sheet.cell(row=row_num, column=4, value=quantity)
                        row_num += 1
            else:
                sheet.cell(row=3, column=1, value="No orders")
                sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=4)
            apply_data_styling(sheet, 3)

        # === Day Total (Admin, Chef, Staff) ===
        if role in ["admin", "chef", "staff"]:
            sheet = workbook.create_sheet(title=f"{day} Total"[:31])
            sheet.sheet_properties.tabColor = "FFD700"
            apply_header_styling(sheet, f"Day Total for {day} ({school})", ["Menu Item", "Total Quantity"])
            row_num = 3
            grand_total = 0
            if day_totals.get(day):
                for menu_name, qty in sorted(day_totals[day].items()):
                    sheet.cell(row=row_num, column=1, value=menu_name)
                    sheet.cell(row=row_num, column=2, value=qty)
                    grand_total += qty
                    row_num += 1
                sheet.cell(row=row_num, column=1, value="Grand Total")
                cell = sheet.cell(row=row_num, column=2, value=grand_total)
                cell.font = Font(bold=True)
                cell.fill = total_fill
            else:
                sheet.cell(row=3, column=1, value="No orders")
                sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=2)
            apply_data_styling(sheet, 3)

        # === Sticker Sheet (Primary Only for Admin) ===
        if school_type == "primary" and role == "admin":
            sheet = workbook.create_sheet(title=f"{day} Stickers"[:31])
            sheet.sheet_properties.tabColor = "92D050"
            apply_header_styling(sheet, f"{day} Sticker Sheet for {school} (Primary)", ["Student Name", "Teacher Name", "Menu Item", "Quantity"])
            row_num = 3
            has_data = False
            for teacher_name, orders in grouped_orders.get(day, {}).items():
                for order_data in orders:
                    student_name = order_data.get("student_name", "Unknown")
                    for menu_name, quantity in order_data["order_items"].items():
                        sheet.cell(row=row_num, column=1, value=student_name)
                        sheet.cell(row=row_num, column=2, value=teacher_name)
                        sheet.cell(row=row_num, column=3, value=menu_name)
                        sheet.cell(row=row_num, column=4, value=quantity)
                        row_num += 1
                        has_data = True
            if not has_data:
                sheet.cell(row=3, column=1, value="No orders")
                sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=4)
            apply_data_styling(sheet, 3)

        # === Detailed Day Total (Admin + Staff) ===
        if role in ["admin", "staff"]:
            sheet = workbook.create_sheet(title=f"{day} Canteen Staff "[:31])
            sheet.sheet_properties.tabColor = "FF9966"
            apply_header_styling(sheet, f"Detailed Day Total for {day} ({school})", ["Teacher/Class", "Student Name", "Menu Item", "Quantity"])
            row_num = 3
            has_data = False

            for entity_name, orders in grouped_orders.get(day, {}).items():
                sorted_orders = sorted(orders, key=lambda x: x.get("student_name", ""))
                for order_data in sorted_orders:
                    student_name = order_data.get("student_name", "Unknown")
                    for menu_name, qty in order_data["order_items"].items():
                        sheet.cell(row=row_num, column=1, value=entity_name)
                        sheet.cell(row=row_num, column=2, value=student_name)
                        sheet.cell(row=row_num, column=3, value=menu_name)
                        sheet.cell(row=row_num, column=4, value=qty)
                        row_num += 1
                        has_data = True

            if not has_data:
                sheet.cell(row=3, column=1, value="No orders")
                sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=4)

            apply_data_styling(sheet, 3)

    if not has_visible_sheet:
        sheet = workbook.create_sheet(title="No Data")
        sheet.cell(row=1, column=1, value="No orders found for the given filters.")

    return workbook

@api_view(['POST'])
def download_menu(request):
    try:
        school_id = request.data.get('school_id')
        school_type = request.data.get('school_type')
        role = request.data.get('role', 'admin')
        day_filter = request.data.get('day')

        if not school_id or not school_type:
            return Response({'error': 'Both school_id and school_type are required.'}, status=status.HTTP_400_BAD_REQUEST)

        if school_type not in ['primary', 'secondary']:
            return Response({'error': 'Invalid school_type.'}, status=status.HTTP_400_BAD_REQUEST)

        school = PrimarySchool.objects.get(id=school_id) if school_type == 'primary' else SecondarySchool.objects.get(id=school_id)
        student_orders, staff_orders = fetch_orders(school_id, school_type, target_day=day_filter)

        if not student_orders.exists() and not staff_orders.exists():
            student_orders, staff_orders = [], []
        # ✅ Generate Excel
        workbook = generate_workbook(
            school, student_orders, staff_orders,
            school_type, role=role, day_filter=day_filter
        )

        # Clean and normalize names
        school_name = (
            school.secondary_school_name if school_type == 'secondary' else school.school_name
        ).replace(' ', '_')

        day_part = f"{day_filter}" if day_filter else "weekly"
        filename = f"{role}_{day_part}_orderSheet_{school_name}.xlsx"

        menu_files_directory = settings.MENU_FILES_ROOT
        os.makedirs(menu_files_directory, exist_ok=True)

        file_path = os.path.join(menu_files_directory, filename)
        workbook.save(file_path)

        return Response({
            'message': 'File generated successfully!',
            'download_link': f"{settings.MENU_FILES_URL}{filename}"
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def download_all_schools_menu(request):
    try:
        role = request.data.get('role')
        day_filter = request.data.get('day')

        if role not in ['admin', 'chef', 'staff']:
            return Response(
                {'error': 'Invalid role. Must be one of: admin, chef, staff.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Create a zip file in memory
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:

            # Process primary schools
            for school in PrimarySchool.objects.all():
                student_orders, staff_orders = fetch_orders(school.id, 'primary')
                workbook = generate_workbook(
                    school, student_orders, staff_orders, 'primary', role=role, day_filter=day_filter
                )

                school_name = school.school_name.replace(' ', '_')
                day_part = day_filter if day_filter else 'weekly'
                filename = f"primary_{school_name}_{day_part}_orders_{role}.xlsx"

                with BytesIO() as excel_buffer:
                    workbook.save(excel_buffer)
                    excel_buffer.seek(0)
                    zip_file.writestr(filename, excel_buffer.getvalue())

            # Process secondary schools
            for school in SecondarySchool.objects.all():
                student_orders, staff_orders = fetch_orders(school.id, 'secondary')
                workbook = generate_workbook(
                    school, student_orders, staff_orders, 'secondary', role=role, day_filter=day_filter
                )

                school_name = school.secondary_school_name.replace(' ', '_')
                day_part = day_filter if day_filter else 'weekly'
                filename = f"secondary_{school_name}_{day_part}_orders_{role}.xlsx"

                with BytesIO() as excel_buffer:
                    workbook.save(excel_buffer)
                    excel_buffer.seek(0)
                    zip_file.writestr(filename, excel_buffer.getvalue())

        zip_buffer.seek(0)

        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        day_part = day_filter if day_filter else 'weekly'
        response['Content-Disposition'] = f'attachment; filename="all_schools_{day_part}_orders_{role}.zip"'
        return response

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


WEEK_DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

def fetch_manager_orders(target_day=None):
    current_time = datetime.now()
    current_week_number = current_time.isocalendar()[1]
    current_year = current_time.year

    cutoff_day = 4  # Friday
    cutoff_hour = 14  # 2PM

    is_past_cutoff = (
        current_time.weekday() > cutoff_day or
        (current_time.weekday() == cutoff_day and current_time.hour >= cutoff_hour)
    )

    target_week = current_week_number + 1 if is_past_cutoff else current_week_number
    target_year = current_year

    if target_week > 52:
        target_week = 1
        target_year += 1

    filter_kwargs = {"week_number": target_week, "year": target_year}
    if target_day:
        filter_kwargs["selected_day__iexact"] = target_day

    return (
        ManagerOrder.objects
        .filter(**filter_kwargs)
        .exclude(status__iexact="cancelled")
        .order_by("selected_day")
    )


@api_view(['GET']) 
def download_manager_orders(request):
    try:
        day_filter = request.GET.get('day')
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:

            days_to_process = [day_filter] if day_filter else WEEK_DAYS

            for day in days_to_process:
                workbook = generate_manager_workbook(day)
                
                day_part = day if day else 'weekly'
                filename = f"manager_orders_{day_part}.xlsx"

                with BytesIO() as excel_buffer:
                    workbook.save(excel_buffer)
                    excel_buffer.seek(0)
                    zip_file.writestr(filename, excel_buffer.getvalue())

        zip_buffer.seek(0)

        # Save to disk and return download link
        menu_files_directory = settings.MENU_FILES_ROOT
        os.makedirs(menu_files_directory, exist_ok=True)

        day_part = day_filter if day_filter else 'weekly'
        filename = f"manager_orders_{day_part}.zip"
        file_path = os.path.join(menu_files_directory, filename)
        
        with open(file_path, 'wb') as f:
            f.write(zip_buffer.getvalue())

        return Response({
            'message': 'Manager orders file generated successfully!',
            'download_link': f"{settings.MENU_FILES_URL}{filename}"
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def generate_manager_workbook(target_day):
    workbook = Workbook()
    workbook.remove(workbook.active)

    # Styles (same as your working code)
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    title_font = Font(bold=True, size=14, color="000000")
    title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    def apply_header_styling(sheet, title_text, columns):
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(columns))
        title_cell = sheet.cell(row=1, column=1, value=title_text)
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center_align

        for col_num, column_title in enumerate(columns, 1):
            cell = sheet.cell(row=2, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
            sheet.column_dimensions[get_column_letter(col_num)].width = 20

    def apply_data_styling(sheet, start_row):
        for row in sheet.iter_rows(min_row=start_row):
            for cell in row:
                cell.border = border
                cell.alignment = left_align if cell.column in [1, 2, 4] else center_align

    # Fetch manager orders for the target day
    manager_orders = fetch_manager_orders(target_day=target_day)
    
    # Group orders by school
    orders_by_school = defaultdict(list)
    day_totals = []

    for order in manager_orders.prefetch_related('items', 'manager'):
        school = None
        school_name = "Unknown School"
        
        # Determine school
        if hasattr(order.manager, 'primary_school') and order.manager.primary_school:
            school = order.manager.primary_school
            school_name = school.school_name
        elif hasattr(order.manager, 'secondary_school') and order.manager.secondary_school:
            school = order.manager.secondary_school
            school_name = school.secondary_school_name
        
        # Process order items
        for item in order.items.all():
            order_data = {
                'manager_name': order.manager.username,
                'item_name': item.item,
                'quantity': item.quantity,
                'remarks': item.remarks or "",
                'production_price': float(item.production_price or 0),
                'total_price': float(item.quantity) * float(item.production_price or 0),
                'order_id': order.id,
                'school_name': school_name
            }
            orders_by_school[school_name].append(order_data)
            day_totals.append(order_data)

    # Create sheets for each school
    for school_name, orders in orders_by_school.items():
        # Clean sheet name (Excel has 31 char limit)
        sheet_name = school_name[:31]
        sheet = workbook.create_sheet(title=sheet_name)
        
        title = f"Manager Orders for {target_day} - {school_name}"
        columns = ["Manager", "Item", "Quantity", "Remarks", "Production Price", "Total Price"]
        apply_header_styling(sheet, title, columns)

        row_num = 3
        for order_data in orders:
            sheet.cell(row=row_num, column=1, value=order_data['manager_name'])
            sheet.cell(row=row_num, column=2, value=order_data['item_name'])
            sheet.cell(row=row_num, column=3, value=order_data['quantity'])
            sheet.cell(row=row_num, column=4, value=order_data['remarks'])
            sheet.cell(row=row_num, column=5, value=order_data['production_price'])
            sheet.cell(row=row_num, column=6, value=order_data['total_price'])
            row_num += 1

        if row_num == 3:
            sheet.cell(row=3, column=1, value="No manager orders")
            sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=6)

        apply_data_styling(sheet, 3)

    # Create Day Total sheet (summary of all schools)
    total_sheet = workbook.create_sheet(title=f"{target_day} Total")
    apply_header_styling(total_sheet, f"Manager Orders Summary - {target_day}", 
                        ["School", "Manager", "Item", "Quantity", "Remarks", "Production Price", "Total Price"])

    row_num = 3
    if not day_totals:
        total_sheet.cell(row=3, column=1, value="No manager orders")
        total_sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=7)
    else:
        for order_data in day_totals:
            total_sheet.cell(row=row_num, column=1, value=order_data['school_name'])
            total_sheet.cell(row=row_num, column=2, value=order_data['manager_name'])
            total_sheet.cell(row=row_num, column=3, value=order_data['item_name'])
            total_sheet.cell(row=row_num, column=4, value=order_data['quantity'])
            total_sheet.cell(row=row_num, column=5, value=order_data['remarks'])
            total_sheet.cell(row=row_num, column=6, value=order_data['production_price'])
            total_sheet.cell(row=row_num, column=7, value=order_data['total_price'])
            row_num += 1

    apply_data_styling(total_sheet, 3)

    return workbook
@api_view(["GET"])
def get_user_count(request):
    try:
      
        parent_count = ParentRegisteration.objects.count()
        staff_count = StaffRegisteration.objects.count()
        secondary_student_count = SecondaryStudent.objects.count()
        primary_student_count = PrimaryStudentsRegister.objects.count()

    
        total_user_count = parent_count + staff_count + secondary_student_count + primary_student_count

        response_data = {
            "total_user_count": total_user_count,
           
        }

        return Response(response_data, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(["GET"])
def get_active_status_menu(request):
    primary_schools = PrimarySchool.objects.all()
    secondary_schools = SecondarySchool.objects.all()

    schools_data = []

    # Handle primary schools
    for school in primary_schools:
        menus = Menu.objects.filter(primary_schools=school)

        # Find active menus for this school
        active_menu = menus.filter(is_active=True).first()

        schools_data.append({
            "school_type": "primary",
            "school_id": school.id,
            "school_name": school.school_name,
            "is_active": bool(active_menu),
            "cycle_name": active_menu.cycle_name if active_menu else None
        })

    # Handle secondary schools
    for school in secondary_schools:
        menus = Menu.objects.filter(secondary_schools=school)

        active_menu = menus.filter(is_active=True).first()

        schools_data.append({
            "school_type": "secondary",
            "school_id": school.id,
            "school_name": school.secondary_school_name,
            "is_active": bool(active_menu),
            "cycle_name": active_menu.cycle_name if active_menu else None
        })

    return Response({"schools": schools_data})



@api_view(['POST'])
def deactivate_menus(request):
    school_id = request.data.get('school_id')
    school_type = request.data.get('school_type')
    cycle_name = request.data.get('cycle_name')

    if not school_id or not school_type or not cycle_name:
        return Response({'error': 'Missing required fields'}, status=status.HTTP_400_BAD_REQUEST)

    school_model = PrimarySchool if school_type == 'primary' else SecondarySchool
    school = school_model.objects.filter(id=school_id).first()
    if not school:
        return Response({'error': f'{school_type.capitalize()} School not found'}, status=status.HTTP_404_NOT_FOUND)

    menus = Menu.objects.filter(cycle_name=cycle_name)
    if not menus.exists():
        return Response({'error': f'No menus found for cycle "{cycle_name}".'}, status=status.HTTP_404_NOT_FOUND)

    for menu in menus:
        if school_type == 'primary':
            menu.primary_schools.remove(school)
        else:
            menu.secondary_schools.remove(school)

        # If no schools left, deactivate menu
        if menu.primary_schools.count() == 0 and menu.secondary_schools.count() == 0:
            menu.is_active = False
            menu.save()

    return Response({'message': f'Cycle "{cycle_name}" deactivated for {school.school_name if school_type == "primary" else school.secondary_school_name}.'}, status=status.HTTP_200_OK)


@api_view(['POST'])
def create_cycle(request):

    cycle_name = request.data.get('cycle_name')
    menu_date = datetime.now().date()

    # Validate cycle name
    if not cycle_name.isalnum() and " " not in cycle_name:
        return Response({'error': 'Cycle Name cannot contain special characters!'}, status=status.HTTP_400_BAD_REQUEST)

    # Check if a menu with the same cycle name already exists
    if Menu.objects.filter(cycle_name=cycle_name).exists():
        return Response({'error': 'A menu with the same cycle name already exists.'}, status=status.HTTP_400_BAD_REQUEST)

    # Define the days
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    created_menus = []

    # Iterate through each day
    for day in days:
        categories = request.data.get(f'category_{day}')
        item_names = request.data.get(f'item_names_{day}')
        prices = request.data.get(f'price_{day}')

        # Skip processing for this day if any of the required fields are None, empty list, or filled with only null or empty values
        if not categories or not item_names or not prices or \
           all(val is None or val == "" for val in categories + item_names + prices):
            continue

        # Check if the data for the day is in the correct format (list of values)
        if not isinstance(categories, list) or not isinstance(item_names, list) or not isinstance(prices, list):
            return Response({'error': f'Invalid data format for {day}. Expecting lists of categories, item names, and prices.'}, status=status.HTTP_400_BAD_REQUEST)

        # Ensure the lengths of the lists are consistent
        if len(categories) != len(item_names) or len(item_names) != len(prices):
            return Response({'error': f'Inconsistent data length for {day}. Categories, item names, and prices must have the same number of items.'}, status=status.HTTP_400_BAD_REQUEST)

        # Process each menu item for the day
        for category_id, menu_name, price in zip(categories, item_names, prices):
            # Skip invalid or incomplete data (empty or null values)
            if category_id is None or not menu_name or price is None or price == "":
                continue

            try:
                price = float(price)
                if price < 0:
                    return Response({'error': f'Price for {menu_name} on {day} must be positive.'}, status=status.HTTP_400_BAD_REQUEST)
            except ValueError:
                return Response({'error': f'Invalid price format for {menu_name} on {day}.'}, status=status.HTTP_400_BAD_REQUEST)

            # Check if the category exists
            try:
                category = Categories.objects.get(id=category_id)
            except Categories.DoesNotExist:
                return Response({'error': f'Category {category_id} not found for {menu_name} on {day}.'}, status=status.HTTP_404_NOT_FOUND)

            # Create and save the menu item
            menu = Menu(
                name=menu_name,
                price=price,
                menu_day=day,
                cycle_name=cycle_name,
                menu_date=menu_date,
                category=category
            )
            menu.save()
            created_menus.append(MenuSerializer(menu).data)

    # Return a success response with the created menus
    return Response({'message': 'Menus created successfully!', 'menus': created_menus}, status=status.HTTP_201_CREATED)


@api_view(['POST', 'PUT', 'DELETE'])
def get_cycle_menus(request):
    cycle_name = request.data.get('cycle_name')

    if not cycle_name:
        return Response({'error': 'Cycle name is required'}, status=status.HTTP_400_BAD_REQUEST)

    # ✅ Fetch Menus
    if request.method == 'POST':
        menus = Menu.objects.filter(cycle_name=cycle_name)
        if not menus.exists():
            return Response({'error': f'No menus found for cycle name: {cycle_name}'}, status=status.HTTP_404_NOT_FOUND)

        serialized_menus = MenuSerializer(menus, many=True)
        return Response({'message': 'Menus fetched successfully!', 'menus': serialized_menus.data}, status=status.HTTP_200_OK)

    # ✅ Update/Replace Menus - SAME FORMAT AS create_cycle
    elif request.method == 'PUT':
        menu_date = datetime.now().date()

        # Delete old menus for this cycle
        Menu.objects.filter(cycle_name=cycle_name).delete()

        # Define the days
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        created_menus = []

        # Iterate through each day
        for day in days:
            categories = request.data.get(f'category_{day}')
            item_names = request.data.get(f'item_names_{day}')
            prices = request.data.get(f'price_{day}')

            # Skip processing for this day if any of the required fields are None, empty list, or filled with only null or empty values
            if not categories or not item_names or not prices or \
               all(val is None or val == "" for val in categories + item_names + prices):
                continue

            # Check if the data for the day is in the correct format (list of values)
            if not isinstance(categories, list) or not isinstance(item_names, list) or not isinstance(prices, list):
                return Response({'error': f'Invalid data format for {day}. Expecting lists of categories, item names, and prices.'}, status=status.HTTP_400_BAD_REQUEST)

            # Ensure the lengths of the lists are consistent
            if len(categories) != len(item_names) or len(item_names) != len(prices):
                return Response({'error': f'Inconsistent data length for {day}. Categories, item names, and prices must have the same number of items.'}, status=status.HTTP_400_BAD_REQUEST)

            # Process each menu item for the day
            for category_data, menu_name, price in zip(categories, item_names, prices):
                # Skip invalid or incomplete data (empty or null values)
                if category_data is None or not menu_name or price is None or price == "":
                    continue

                try:
                    price = float(price)
                    if price < 0:
                        return Response({'error': f'Price for {menu_name} on {day} must be positive.'}, status=status.HTTP_400_BAD_REQUEST)
                except ValueError:
                    return Response({'error': f'Invalid price format for {menu_name} on {day}.'}, status=status.HTTP_400_BAD_REQUEST)

                # Handle category - it could be ID (integer) or name (string)
                if isinstance(category_data, str):
                    # Try to find category by name
                    try:
                        category = Categories.objects.get(name_category=category_data)
                    except Categories.DoesNotExist:
                        return Response({'error': f'Category "{category_data}" not found for {menu_name} on {day}.'}, status=status.HTTP_404_NOT_FOUND)
                else:
                    # Assume it's a category ID
                    try:
                        category = Categories.objects.get(id=category_data)
                    except Categories.DoesNotExist:
                        return Response({'error': f'Category ID {category_data} not found for {menu_name} on {day}.'}, status=status.HTTP_404_NOT_FOUND)

                # Create and save the menu item
                menu = Menu(
                    name=menu_name,
                    price=price,
                    menu_day=day,
                    cycle_name=cycle_name,
                    menu_date=menu_date,
                    category=category
                )
                menu.save()
                created_menus.append(MenuSerializer(menu).data)

        # Return a success response with the created menus
        return Response({'message': 'Menus updated successfully!', 'menus': created_menus}, status=status.HTTP_200_OK)

    # ✅ Delete Menus
    elif request.method == 'DELETE':
        menus = Menu.objects.filter(cycle_name=cycle_name)
        if not menus.exists():
            return Response({'error': f'No menus found for cycle name: {cycle_name}'}, status=status.HTTP_404_NOT_FOUND)

        deleted_count, _ = menus.delete()
        return Response({'message': f'{deleted_count} menus deleted successfully.'}, status=status.HTTP_200_OK)
@api_view(['GET'])
def get_all_cycles_with_menus(request):
    cycle_names = Menu.objects.values_list('cycle_name', flat=True).distinct()

    if not cycle_names:
        return Response({'error': 'No cycle names found'}, status=status.HTTP_404_NOT_FOUND)

    cycles_with_menus = []

    for cycle_name in cycle_names:
        # Fetch all menus for this cycle
        menus = (
            Menu.objects.filter(cycle_name=cycle_name)
            .order_by('id')
            .values(
                'id',
                'cycle_name',
                'name',
                'category',
                'menu_date',
                'menu_day',
                'price',
                'primary_schools__id',
                'primary_schools__school_name',
                'secondary_schools__id',
                'secondary_schools__secondary_school_name'
            )
        )

        serialized_menus = list(menus)

        cycles_with_menus.append({
            'cycle_name': cycle_name,
            'menus': serialized_menus
        })

    return Response(
        {
            'message': 'Cycles and their menus fetched successfully!',
            'cycles': cycles_with_menus
        },
        status=status.HTTP_200_OK
    )



@api_view(['POST'])
def duplicate_cycle(request):
    original_cycle_name = request.data.get("cycle_name")
    new_cycle_name = request.data.get("new_cycle_name")

    if not original_cycle_name or not new_cycle_name:
        return Response(
            {"error": "Both 'cycle_name' and 'new_cycle_name' are required."},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Check if original cycle exists
    menus = Menu.objects.filter(cycle_name=original_cycle_name)
    if not menus.exists():
        return Response(
            {"error": f"No menus found for cycle '{original_cycle_name}'"},
            status=status.HTTP_404_NOT_FOUND
        )

    # Prevent overwriting an existing cycle
    if Menu.objects.filter(cycle_name=new_cycle_name).exists():
        return Response(
            {"error": f"A cycle with name '{new_cycle_name}' already exists."},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        with transaction.atomic():
            new_menus = []
            for menu in menus:
                menu.pk = None  # Reset PK so Django creates a new record
                menu.cycle_name = new_cycle_name
                new_menus.append(menu)
            Menu.objects.bulk_create(new_menus)

        return Response(
            {"message": f"Cycle '{original_cycle_name}' duplicated as '{new_cycle_name}' successfully!"},
            status=status.HTTP_201_CREATED
        )
    except Exception as e:
        return Response(
            {"error": f"Something went wrong: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET', 'DELETE'])
def all_users_report(request):
    if request.method == 'GET':
        data = []

        for parent in ParentRegisteration.objects.all():
            data.append({"id": parent.id, "email": parent.email, "type": "parent"})

        for student in SecondaryStudent.objects.all():
            data.append({"id": student.id, "email": student.email, "type": "student"})

        for staff in StaffRegisteration.objects.all():
            data.append({"id": staff.id, "email": staff.email, "type": "staff"})

        return Response(data)

    elif request.method == 'DELETE':
        ParentRegisteration.objects.all().delete()
        SecondaryStudent.objects.all().delete()
        StaffRegisteration.objects.all().delete()
        return Response({"message": "All users deleted successfully"})


from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from rest_framework_simplejwt.tokens import RefreshToken
import jwt
from admin_section.models import (
    UnverifiedUser,
    ParentRegisteration,
    StaffRegisteration,
    SecondaryStudent
)

def generate_login_response(user, user_type):
    refresh = RefreshToken.for_user(user)
    return Response({
        "access": str(refresh.access_token),
        "refresh": str(refresh),
        "user_type": user_type,
        "user_id": user.id,
        "message": "Login successful"
    }, status=200)

@api_view(["POST"])
def social_callback_register(request):
    token = request.data.get("access_token")
    provider = request.data.get("provider")
    email = None
    first_name = ""
    last_name = ""

    # ----------------- GOOGLE AUTHENTICATION ----------------- #
    if provider == "google":
        try:
            # Try to decode ID token first
            decoded = jwt.decode(token, options={"verify_signature": False})
            email = decoded.get("email")
            first_name = decoded.get("given_name", "")
            last_name = decoded.get("family_name", "")
            
            # Fallback to userinfo endpoint if names not in ID token
            if not first_name or not last_name:
                user_info = requests.get(
                    "https://www.googleapis.com/oauth2/v3/userinfo",
                    headers={"Authorization": f"Bearer {token}"}
                ).json()
                first_name = user_info.get("given_name", first_name)
                last_name = user_info.get("family_name", last_name)
                
        except Exception as e:
            return Response({"error": f"Google authentication failed: {str(e)}"}, status=400)

    # ----------------- MICROSOFT AUTHENTICATION ----------------- #
    elif provider == "microsoft":
        try:
            resp = requests.get(
                "https://graph.microsoft.com/v1.0/me?" + 
                "$select=mail,userPrincipalName,givenName,surname",
                headers={"Authorization": f"Bearer {token}"}
            )
            if resp.status_code != 200:
                return Response({"error": "Microsoft authentication failed"}, status=400)
                
            user_info = resp.json()
            email = user_info.get("mail") or user_info.get("userPrincipalName")
            first_name = user_info.get("givenName", "")
            last_name = user_info.get("surname", "")
        except Exception as e:
            return Response({"error": f"Microsoft authentication failed: {str(e)}"}, status=400)

    # ----------------- APPLE AUTHENTICATION ----------------- #
    elif provider == "apple":
        try:
            decoded = verify_apple_token(token)
            email = decoded.get("email")
            if not email:
                email = f"{decoded['sub']}@privaterelay.appleid.com"
            first_name = decoded.get("first_name", "")
            last_name = decoded.get("last_name", "")
        except Exception as e:
            return Response({"error": f"Apple authentication failed: {str(e)}"}, status=400)

    # ----------------- COMMON VALIDATION ----------------- #
    if not email:
        return Response({"error": "Could not retrieve email from provider"}, status=400)

    # Check if user already exists
    parent = ParentRegisteration.objects.filter(email=email).first()
    if parent:
        return generate_login_response(parent, "parent")

    staff = StaffRegisteration.objects.filter(email=email).first()
    if staff:
        return generate_login_response(staff, "staff")

    student = SecondaryStudent.objects.filter(email=email).first()
    if student:
        return generate_login_response(student, "student")

    # Create or update unverified user
    user_data = {
        "first_name": first_name or "User",  # Fallback if empty
        "last_name": last_name or "Unknown",
        "provider": provider
    }

    unverified, created = UnverifiedUser.objects.update_or_create(
        email=email,
        defaults={
            "token": uuid.uuid4(),
            "login_method": provider,
            "data": user_data
        }
    )

    return Response({
        "message": "Additional information required to complete registration",
        "token": str(unverified.token),
        "provider": provider,
        "email": email,
        "first_name": first_name,
        "last_name": last_name
    }, status=200)


@api_view(["POST"])
def complete_social_signup(request):
    token = request.data.get("token")
    role = request.data.get("role")  
    data = request.data.get("data")

    if not token or not role or not data:
        return Response({"error": "Missing token, role, or data."}, status=400)

    # Acceptable login methods
    unverified = UnverifiedUser.objects.filter(
        token=token, login_method__in=["google", "microsoft", "apple"]
    ).first()

    if not unverified:
        return Response({"error": "Invalid or expired token"}, status=400)

    # Extract names from unverified data
    first_name = unverified.data.get("first_name", "User")
    last_name = unverified.data.get("last_name", "Unknown")
    email = unverified.email
    phone_no = data.get("phone_no")
    if phone_no:
        # Remove all spaces
        phone_no = phone_no.replace(" ", "").strip()


    # Convert allergy names to IDs
    allergy_names = data.pop("allergies", [])
    allergy_ids = list(
        Allergens.objects.filter(allergy__in=allergy_names).values_list("id", flat=True)
    )

    # Clean up unnecessary fields
    data.pop("user_type", None)
    data.pop("schoolName", None)
    data.pop("password", None)
    data.pop("retypePassword", None)

    created_user = None

    if role == "parent":
        created_user = ParentRegisteration.objects.create(
            first_name=first_name,
            last_name=last_name,
            username=data.get("username", ""),
            email=email,
            phone_no=phone_no,   
            password="social_dummy_password"
        )

    elif role == "student":
        school_id = data.get("school_id")
        school = SecondarySchool.objects.filter(id=school_id).first()
        created_user = SecondaryStudent.objects.create(
            first_name=first_name,
            last_name=last_name,
            username=data.get("username", ""),
            email=email,
            phone_no=phone_no,
            class_year=data.get("class_year", ""),
            school=school,
            password="social_dummy_password"
        )

    elif role == "staff":
        school_type = data.get("school_type", "").lower()
        school_id = data.get("school_id")

        primary_school = None
        secondary_school = None

        if school_type == "primary":
            primary_school = PrimarySchool.objects.filter(id=school_id).first()
        elif school_type == "secondary":
            secondary_school = SecondarySchool.objects.filter(id=school_id).first()

        created_user = StaffRegisteration.objects.create(
            first_name=first_name,
            last_name=last_name,
            username=data.get("username", ""),
            email=email,
            phone_no=phone_no,
            password="social_dummy_password",
            primary_school=primary_school,
            secondary_school=secondary_school,
        )

    # Set allergies if applicable
    if created_user and allergy_ids:
        created_user.allergies.set(allergy_ids)

    # Delete unverified user now that account is created
    unverified.delete()

    # Generate JWT tokens
    refresh = RefreshToken.for_user(created_user)
    access_token = str(refresh.access_token)
    refresh_token = str(refresh)

    return Response({
        "access": access_token,
        "refresh": refresh_token,
        "user_type": role,  # Directly use role as user_type
        "user_id": created_user.id,
        "message": "Signup and login successful"
    }, status=status.HTTP_200_OK)

@api_view(["GET"])
def get_app_version(request, platform):
    try:
        version = AppVersion.objects.get(platform=platform)
        serializer = AppVersionSerializer(version)
        return Response(serializer.data)
    except AppVersion.DoesNotExist:
        return Response({"error": "Platform not found"}, status=404)
@api_view(['POST'])
def make_menu_available(request):
    menu_id = request.data.get('menu_id')

    if not menu_id:
        return Response({'error': 'menu_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        menu_item = MenuItems.objects.get(id=menu_id)
    except MenuItems.DoesNotExist:
        return Response({'error': 'Menu item not found'}, status=status.HTTP_404_NOT_FOUND)

    menu_item.is_available = True
    menu_item.save()

    return Response({
        'message': f'Menu item "{menu_item.item_name}" is now available.',
        'menu_id': menu_item.id,
        'is_available': menu_item.is_available,
    }, status=status.HTTP_200_OK)


@api_view(['POST'])
def make_menu_unavailable(request):
    menu_id = request.data.get('menu_id')

    if not menu_id:
        return Response({'error': 'menu_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        menu_item = MenuItems.objects.get(id=menu_id)
    except MenuItems.DoesNotExist:
        return Response({'error': 'Menu item not found'}, status=status.HTTP_404_NOT_FOUND)

    menu_item.is_available = False
    menu_item.save()

    return Response({
        'message': f'Menu item "{menu_item.item_name}" is now unavailable.',
        'menu_id': menu_item.id,
        'is_available': menu_item.is_available,
    }, status=status.HTTP_200_OK)



@api_view(['POST'])
def manager_login(request):
    username = request.data.get("username", "").strip().lower()
    password = request.data.get("password")
    print('username',username)
    print('password',password)

    if not username or not password:
        return Response({'detail': 'Both username and password are required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        manager = Manager.objects.get(username=username)
    except Manager.DoesNotExist:
        return Response({'detail': 'Invalid username or password.'}, status=status.HTTP_401_UNAUTHORIZED)

    if not check_password(password, manager.password):
        return Response({'detail': 'Invalid username or password.'}, status=status.HTTP_401_UNAUTHORIZED)

    data = {
        "id": manager.id,
        "username": manager.username,
        "school_type": manager.school_type,
        "primary_school": manager.primary_school.id if manager.primary_school else None,
        "secondary_school": manager.secondary_school.id if manager.secondary_school else None,
    }

    return Response({"detail": "Login successful!", "manager": data}, status=status.HTTP_200_OK)



@api_view(['POST'])
def worker_login(request):
    username = request.data.get("username", "").strip().lower()
    password = request.data.get("password")

    if not username or not password:
        return Response({'detail': 'Both username and password are required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        worker = Worker.objects.get(username=username)
    except Worker.DoesNotExist:
        return Response({'detail': 'Invalid username or password.'}, status=status.HTTP_401_UNAUTHORIZED)

    if not check_password(password, worker.password):
        return Response({'detail': 'Invalid username or password.'}, status=status.HTTP_401_UNAUTHORIZED)

    data = {
        "id": worker.id,
        "username": worker.username,
    }

    return Response({"detail": "Login successful!", "worker": data}, status=status.HTTP_200_OK)


class CreateManagerOrderAPIView(APIView):
    """
    Create an order for Managers.
    This view takes:
      - manager_id
      - cart (list of {day, item, quantity, remarks})
    """

    def post(self, request, *args, **kwargs):
        try:
            data = request.data
            manager_id = data.get("manager_id")
            cart_items = data.get("cart", [])

            if not manager_id:
                return Response({"error": "Manager ID is required."}, status=status.HTTP_400_BAD_REQUEST)

            if not cart_items:
                return Response({"error": "Cart cannot be empty."}, status=status.HTTP_400_BAD_REQUEST)

            manager = Manager.objects.filter(id=manager_id).first()
            if not manager:
                return Response({"error": "Manager not found."}, status=status.HTTP_404_NOT_FOUND)

            grouped_items = {}
            for item in cart_items:
                day = item.get("day")
                if not day:
                    return Response({"error": "Each item must include a day."}, status=status.HTTP_400_BAD_REQUEST)
                grouped_items.setdefault(day, []).append(item)

            created_orders = []
            with transaction.atomic():
                for day, items in grouped_items.items():
                    today = datetime.now()

                    try:
                        target_day_index = [
                            "monday", "tuesday", "wednesday", "thursday",
                            "friday", "saturday", "sunday"
                        ].index(day.lower())
                    except ValueError:
                        return Response({"error": f"Invalid day '{day}'."}, status=status.HTTP_400_BAD_REQUEST)

                    # ✅ Same calculation as user orders
                    days_ahead = (target_day_index - today.weekday() + 7) % 7
                    order_date = today + timedelta(days=days_ahead)
                    week_number = order_date.isocalendar()[1]
                    year = order_date.year

                    # ✅ Create ManagerOrder with correct order_date
                    order = ManagerOrder.objects.create(
                        manager=manager,
                        week_number=week_number,
                        year=year,
                        selected_day=day,
                        order_date=order_date.date(),  # <-- Added line
                        status="pending",
                        is_delivered=False,
                    )

                    for cart_item in items:
                        item_name = cart_item.get("item")
                        quantity = int(cart_item.get("quantity", 1))
                        remarks = cart_item.get("remarks", "")

                        menu_item = MenuItems.objects.filter(item_name__iexact=item_name).first()
                        if not menu_item:
                            return Response(
                                {"error": f"Menu item '{item_name}' not found."},
                                status=status.HTTP_404_NOT_FOUND
                            )

                        ManagerOrderItem.objects.create(
                            order=order,
                            day=day,
                            item=item_name,
                            quantity=quantity,
                            remarks=remarks,
                            menu_item=menu_item
                        )

                    created_orders.append(order)

            serializer = ManagerOrderSerializer(created_orders, many=True)
            orders_data = serializer.data

            for order in orders_data:
                order["order_id"] = f"m_{order['id']}"
                del order["id"]

            return Response({
                "message": "Manager orders created successfully.",
                "orders": orders_data
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




# ------------------------------
# Create Document (Admin only)
# ------------------------------
@api_view(['POST'])
def create_document(request):
    title = request.data.get('title')
    content = request.data.get('content')

    if not title or not content:
        return Response({'error': 'Title and content are required.'}, status=status.HTTP_400_BAD_REQUEST)


    document = Document.objects.create(
        title=title,
        content=content,
    )

    return Response({
        'message': 'Document created successfully.',
        'document_id': document.id,
        'title': document.title,
        'content': document.content,
    }, status=status.HTTP_201_CREATED)


# ------------------------------
# Fetch All Documents
# ------------------------------
@api_view(['GET'])
def get_all_documents(request):
    documents = Document.objects.all().order_by('-created_at')
    data = [{
        'id': doc.id,
        'title': doc.title,
        'content': doc.content,
        'created_at': doc.created_at,
        'updated_at': doc.updated_at
    } for doc in documents]
    return Response(data, status=status.HTTP_200_OK)


# ------------------------------
# Edit Document
# ------------------------------
@api_view(['POST'])
def edit_document(request):
    document_id = request.data.get('document_id')
    title = request.data.get('title')
    content = request.data.get('content')

    if not document_id:
        return Response({'error': 'document_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        document = Document.objects.get(id=document_id)
    except Document.DoesNotExist:
        return Response({'error': 'Document not found.'}, status=status.HTTP_404_NOT_FOUND)

    if title:
        document.title = title
    if content:
        document.content = content

    document.save()
    return Response({'message': 'Document updated successfully.'}, status=status.HTTP_200_OK)


# ------------------------------
# Delete Document
# ------------------------------
@api_view(['POST'])
def delete_document(request):
    document_id = request.data.get('document_id')
    if not document_id:
        return Response({'error': 'document_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        document = Document.objects.get(id=document_id)
        document.delete()
    except Document.DoesNotExist:
        return Response({'error': 'Document not found.'}, status=status.HTTP_404_NOT_FOUND)

    return Response({'message': 'Document deleted successfully.'}, status=status.HTTP_200_OK)


# ------------------------------
# Fetch Documents for a Worker (with read/unread status)
# ------------------------------
@api_view(['POST'])
def get_worker_documents(request):
    worker_id = request.data.get('worker_id')

    if not worker_id:
        return Response({'error': 'worker_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        worker = Worker.objects.get(id=worker_id)
    except Worker.DoesNotExist:
        return Response({'error': 'Worker not found.'}, status=status.HTTP_404_NOT_FOUND)

    documents = Document.objects.all()
    data = []

    for doc in documents:
        status_obj = WorkerDocumentStatus.objects.filter(worker=worker, document=doc).first()
        data.append({
            'document_id': doc.id,
            'title': doc.title,
            'status': status_obj.status if status_obj else 'unread',
            'read_at': status_obj.read_at if status_obj else None
        })

    return Response(data, status=status.HTTP_200_OK)


# ------------------------------
# Fetch Document Detail (by document_id)
# ------------------------------
@api_view(['POST'])
def get_document_detail(request):
    document_id = request.data.get('document_id')

    if not document_id:
        return Response({'error': 'document_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        document = Document.objects.get(id=document_id)
    except Document.DoesNotExist:
        return Response({'error': 'Document not found.'}, status=status.HTTP_404_NOT_FOUND)

    return Response({
        'id': document.id,
        'title': document.title,
        'content': document.content,
        'created_at': document.created_at,
        'updated_at': document.updated_at,
    }, status=status.HTTP_200_OK)


# ------------------------------
# Mark Document as Read (worker action)
# ------------------------------
@api_view(['POST'])
def mark_document_read(request):
    worker_id = request.data.get('worker_id')
    document_id = request.data.get('document_id')
    status_value = request.data.get('status')  # 'read' or 'unread'

    if not worker_id or not document_id or not status_value:
        return Response({'error': 'worker_id, document_id and status are required.'},
                        status=status.HTTP_400_BAD_REQUEST)

    try:
        worker = Worker.objects.get(id=worker_id)
        document = Document.objects.get(id=document_id)
    except (Worker.DoesNotExist, Document.DoesNotExist):
        return Response({'error': 'Worker or Document not found.'}, status=status.HTTP_404_NOT_FOUND)

    obj, created = WorkerDocumentStatus.objects.get_or_create(worker=worker, document=document)
    obj.status = status_value
    obj.read_at = timezone.now() if status_value == 'read' else None
    obj.save()

    return Response({
        'message': f'Document marked as {status_value}.',
        'worker': worker.username,
        'document': document.title,
        'read_at': obj.read_at
    }, status=status.HTTP_200_OK)



@api_view(['GET'])
def export_worker_document_status(request):
    """
    Export Excel showing all documents vs all workers with read status.
    Rows = Documents, Columns = Workers, Values = Read Timestamp or 'Unread'
    """
    try:
        # Create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Document Status"

        # Styles
        header_fill = PatternFill(start_color="009c5b", end_color="009c5b", fill_type="solid")
        read_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  
        unread_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") 
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        workers = list(Worker.objects.all())
        documents = list(Document.objects.all())

        # Header row
        headers = ["Document Title"] + [worker.username for worker in workers]
        ws.append(headers)

        # Apply header styles
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Fill data rows
        for doc in documents:
            row_data = [doc.title]
            for worker in workers:
                status_obj = WorkerDocumentStatus.objects.filter(worker=worker, document=doc).first()

                if status_obj and status_obj.status == "read":
                    value = status_obj.read_at.strftime("%Y-%m-%d %H:%M") if status_obj.read_at else "Read"
                else:
                    value = "Unread"

                row_data.append(value)

            ws.append(row_data)

        # Apply coloring for each cell based on status
        for row_idx in range(2, len(documents) + 2):
            for col_idx in range(2, len(workers) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value == "Unread":
                    cell.fill = unread_fill
                else:
                    cell.fill = read_fill
                cell.alignment = center_align

        # Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 4

        # Save workbook to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Prepare response
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="document_status_report.xlsx"'
        return response

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
